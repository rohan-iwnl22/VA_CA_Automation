"""Tests for the excel_writer modules."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from va_ca_automation.excel_writer.data_writer import (
    DATA_FONT,
    TEMPLATE_COLUMNS,
    write_va_data_rows,
    write_va_report_header,
)
from va_ca_automation.excel_writer.summary_builder import (
    build_risk_summary,
    build_scope_table,
)
from va_ca_automation.logging.pipeline_logger import PipelineLogger
from va_ca_automation.metadata.engagement_metadata import EngagementMetadata, HostMetadata
from datetime import date


class TestDataWriter:
    def test_write_va_report_header(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "VA Report"

        metadata = EngagementMetadata(
            client_name="Test Client",
            security_tester="Tester",
            reviewed_by="Reviewer",
            report_date=date(2026, 1, 15),
            report_version="1.2",
            scanner_name="Nessus ",
        )
        write_va_report_header(ws, metadata)

        assert ws["C5"].value == "Test Client"
        assert ws["C6"].value == "Tester"
        assert ws["C7"].value == "Reviewer"
        assert ws["C8"].value == date(2026, 1, 15)
        assert ws["C9"].value == "1.2"
        assert ws["C10"].value == "Nessus "
        wb.close()

    def test_write_va_data_rows(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "VA Report"

        df = pd.DataFrame(
            {
                "Sr. no": [1, 2],
                "Vulnerbility Title": ["Vuln A", "Vuln B"],
                "Description": ["Desc A", "Desc B"],
                "Risk": ["Critical", "High"],
                "Host": ["10.0.0.1", "10.0.0.2"],
                "Port": [443, 80],
                "Recommendation ": ["Fix A", "Fix B"],
                "Reference": ["http://ref.com", None],
                "CVE": ["CVE-2024-0001", None],
            }
        )

        last_row = write_va_data_rows(ws, df)

        assert last_row == 15
        assert ws.cell(row=14, column=1).value == 1
        assert ws.cell(row=14, column=2).value == "Vuln A"
        assert ws.cell(row=14, column=4).value == "Critical"
        assert ws.cell(row=15, column=2).value == "Vuln B"
        assert ws.cell(row=15, column=8).value == "N/A"  # Reference null → N/A
        assert ws.cell(row=14, column=1).font.name == "Cambria"
        wb.close()

    def test_write_empty_df(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "VA Report"

        df = pd.DataFrame(columns=TEMPLATE_COLUMNS)
        last_row = write_va_data_rows(ws, df)
        assert last_row == 13  # no data rows
        wb.close()


class TestSummaryBuilder:
    def test_build_scope_table(self):
        va_df = pd.DataFrame({"Host": ["10.0.0.1", "10.0.0.2", "10.0.0.1"]})
        metadata = EngagementMetadata(
            client_name="Test",
            security_tester="T",
            reviewed_by="R",
            report_date=date(2026, 1, 1),
            report_version="1.0",
            host_metadata={
                "10.0.0.1": HostMetadata("10.0.0.1", "Authenticated", "Server"),
            },
        )
        scope = build_scope_table(va_df, metadata)
        assert len(scope) == 2  # 2 distinct hosts
        assert scope.iloc[0]["IP Address"] == "10.0.0.1"
        assert scope.iloc[0]["Scan Type"] == "Authenticated"

    def test_build_risk_summary(self):
        va_df = pd.DataFrame({"Risk": ["Critical"] * 21 + ["High"] * 32 + ["Medium"] * 77 + ["Low"] * 15})
        summary = build_risk_summary(va_df)
        assert summary["Critical"] == 21
        assert summary["High"] == 32
        assert summary["Medium"] == 77
        assert summary["Low"] == 15
        assert summary["Grand Total"] == 145
