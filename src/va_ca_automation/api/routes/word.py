"""POST /api/word — Generate Word report from VA/CA Excel reports."""

from __future__ import annotations

import logging
import tempfile
from pathlib import Path

import pandas as pd
from datetime import date, datetime
from fastapi import APIRouter, Depends, File, Form, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

from ...metadata.engagement_metadata import EngagementMetadata
from ...naming.filename_builder import _sanitize_filename
from ...word_writer.word_report_builder import build_word_report
from ..deps import get_current_user
from ..temp_registry import create_session, store_file

router = APIRouter(tags=["word"])
logger = logging.getLogger("va_ca_automation")


@router.post("/word/read-metadata")
async def read_excel_metadata(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Read client name and other metadata from an uploaded Excel file."""
    temp_dir = Path(tempfile.mkdtemp())
    try:
        content = await file.read()
        suffix = Path(file.filename).suffix if file.filename else ".xlsx"
        tmp_path = temp_dir / (file.filename or f"upload{suffix}")
        tmp_path.write_bytes(content)

        meta = {}
        if _is_va_report(tmp_path):
            meta = _read_va_metadata(tmp_path)
        elif _is_ca_report(tmp_path):
            meta = _read_ca_metadata(tmp_path)

        return JSONResponse(content=meta)
    except Exception as e:
        logger.warning("Could not read metadata: %s", e)
        return JSONResponse(content={})
    finally:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)


def _is_va_report(path: Path) -> bool:
    """Check if an Excel file is a VA report by looking for 'VA Report' sheet."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        has_sheet = "VA Report" in wb.sheetnames
        wb.close()
        return has_sheet
    except Exception:
        return False


def _is_ca_report(path: Path) -> bool:
    """Check if an Excel file is a CA report by looking for 'CA_Report' sheet."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        has_sheet = "CA_Report" in wb.sheetnames
        wb.close()
        return has_sheet
    except Exception:
        return False


def _read_va_metadata(va_path: Path) -> dict:
    """Read engagement metadata from VA Excel report header (rows 5-10, column C)."""
    meta = {}
    try:
        wb = load_workbook(va_path, read_only=True, data_only=True)
        ws = wb["VA Report"]
        meta["client_name"] = ws["C5"].value or ""
        meta["security_tester"] = ws["C6"].value or ""
        meta["reviewed_by"] = ws["C7"].value or ""
        raw_date = ws["C8"].value
        if raw_date:
            if hasattr(raw_date, "strftime"):
                meta["report_date"] = raw_date.strftime("%Y-%m-%d")
            else:
                meta["report_date"] = str(raw_date)
        else:
            meta["report_date"] = ""
        meta["report_version"] = ws["C9"].value or ""
        wb.close()
    except Exception as e:
        logger.warning("Could not read metadata from VA Excel: %s", e)
    return meta


def _read_ca_metadata(ca_path: Path) -> dict:
    """Read engagement metadata from CA Excel report header (rows 5-10, column C)."""
    meta = {}
    try:
        wb = load_workbook(ca_path, read_only=True, data_only=True)
        ws = wb["CA_Report"]
        meta["client_name"] = ws["C5"].value or ""
        meta["security_tester"] = ws["C6"].value or ""
        meta["reviewed_by"] = ws["C7"].value or ""
        raw_date = ws["C8"].value
        if raw_date:
            if hasattr(raw_date, "strftime"):
                meta["report_date"] = raw_date.strftime("%Y-%m-%d")
            else:
                meta["report_date"] = str(raw_date)
        else:
            meta["report_date"] = ""
        meta["report_version"] = ws["C9"].value or ""
        wb.close()
    except Exception as e:
        logger.warning("Could not read metadata from CA Excel: %s", e)
    return meta


def _read_va_data(va_path: Path) -> tuple[pd.DataFrame, dict[str, int]]:
    """Read VA data and risk summary from a generated VA Excel report."""
    va_word_df = pd.DataFrame()
    va_risk_summary = {
        "Critical": 0,
        "High": 0,
        "Medium": 0,
        "Low": 0,
        "Grand Total": 0,
    }
    try:
        va_wb = load_workbook(va_path, read_only=True, data_only=True)
        va_ws = va_wb["VA Report"]
        va_data = []
        for row in va_ws.iter_rows(min_row=14, max_col=9, values_only=True):
            if row[0] is not None:
                va_data.append(row)
        if va_data:
            va_cols = [
                "Sr. no",
                "Vulnerbility Title",
                "Description",
                "Risk",
                "Host",
                "Port",
                "Recommendation ",
                "Reference",
                "CVE",
            ]
            va_word_df = pd.DataFrame(va_data, columns=va_cols)
        va_wb.close()

        va_summary_wb = load_workbook(va_path, read_only=True, data_only=True)
        va_summary_ws = va_summary_wb["Summary"]
        for row in va_summary_ws.iter_rows(
            min_row=16, max_row=20, min_col=5, max_col=6, values_only=True
        ):
            if row[0] and row[1] is not None:
                label = str(row[0]).strip()
                count = int(row[1]) if row[1] else 0
                if label in va_risk_summary:
                    va_risk_summary[label] = count
                elif label == "Grand Total":
                    va_risk_summary["Grand Total"] = count
        va_summary_wb.close()
    except Exception as e:
        logger.warning("Could not read VA Excel for Word report: %s", e)

    return va_word_df, va_risk_summary


def _read_ca_data(ca_path: Path) -> tuple[pd.DataFrame, dict[str, int]]:
    """Read CA data and risk summary from a generated CA Excel report."""
    ca_word_df = pd.DataFrame()
    ca_risk_summary = {"FAILED": 0, "WARNING": 0, "Grand Total": 0}
    try:
        ca_wb = load_workbook(ca_path, read_only=True, data_only=True)
        ca_ws = ca_wb["CA_Report"]
        ca_data = []
        for row in ca_ws.iter_rows(min_row=14, max_col=6, values_only=True):
            if row[0] is not None:
                ca_data.append(row)
        if ca_data:
            ca_cols = [
                "Sr.No.",
                "Title",
                "Host",
                "Description",
                "Solution",
                "Risk",
            ]
            ca_word_df = pd.DataFrame(ca_data, columns=ca_cols)
        ca_wb.close()

        ca_summary_wb = load_workbook(ca_path, read_only=True, data_only=True)
        ca_summary_ws = ca_summary_wb["Summary"]
        for row in ca_summary_ws.iter_rows(
            min_row=16, max_row=18, min_col=5, max_col=6, values_only=True
        ):
            if row[0] and row[1] is not None:
                label = str(row[0]).strip()
                count = int(row[1]) if row[1] else 0
                if label in ca_risk_summary:
                    ca_risk_summary[label] = count
                elif label == "Grand Total":
                    ca_risk_summary["Grand Total"] = count
        ca_summary_wb.close()
    except Exception as e:
        logger.warning("Could not read CA Excel for Word report: %s", e)

    return ca_word_df, ca_risk_summary


@router.post("/word")
async def generate_word(
    files: list[UploadFile] = File(...),
    client_name: str = Form(""),
    client_short_name: str = Form(""),
    security_tester: str = Form(""),
    reviewed_by: str = Form(""),
    device_type: str = Form(""),
    scope: str = Form("Server"),
    phase: str = Form("First"),
    report_type: str = Form("First"),
    report_number: str = Form("1.0"),
    report_date: str = Form(""),
    assessment_start_date: str = Form(""),
    assessment_finish_date: str = Form(""),
    final_retesting_start: str = Form(""),
    final_retesting_finish: str = Form(""),
    released_date: str = Form(""),
    spokesperson_name: str = Form(""),
    spokesperson_designation: str = Form(""),
    spokesperson_email: str = Form(""),
    senior_name: str = Form(""),
    approved_by: str = Form("Mr. Vijay Sawant"),
    # New fields
    report_release_date: str = Form(""),
    period: str = Form(""),
    document_id: str = Form(""),
    change_history_version: str = Form(""),
    change_history_date: str = Form(""),
    change_history_remarks: str = Form(""),
    distribution_name: str = Form(""),
    distribution_organization: str = Form(""),
    distribution_designation: str = Form(""),
    distribution_email: str = Form(""),
    # Auditing Team - Row 1
    auditor_1_name: str = Form(""),
    auditor_1_designation: str = Form(""),
    auditor_1_email: str = Form(""),
    auditor_1_qualifications: str = Form(""),
    auditor_1_cert_in: str = Form("Yes"),
    # Auditing Team - Row 2
    auditor_2_name: str = Form(""),
    auditor_2_designation: str = Form(""),
    auditor_2_email: str = Form(""),
    auditor_2_qualifications: str = Form(""),
    auditor_2_cert_in: str = Form("Yes"),
    current_user: dict = Depends(get_current_user),
):
    """Generate Word report from uploaded VA/CA Excel reports. Returns download URL."""
    temp_dir = Path(tempfile.mkdtemp())
    saved_files = []

    try:
        for upload in files:
            content = await upload.read()
            suffix = Path(upload.filename).suffix if upload.filename else ".xlsx"
            tmp_path = temp_dir / (upload.filename or f"upload_{len(saved_files)}{suffix}")
            tmp_path.write_bytes(content)
            saved_files.append(tmp_path)

        va_path = None
        ca_path = None
        for p in saved_files:
            if va_path is None and _is_va_report(p):
                va_path = p
            elif ca_path is None and _is_ca_report(p):
                ca_path = p

        if va_path is None and ca_path is None:
            return JSONResponse(
                status_code=400,
                content={"detail": "No valid VA or CA report files found. Upload files with 'VA Report' or 'CA_Report' sheets."},
            )

        excel_meta = _read_va_metadata(va_path) if va_path else {}

        if not client_name and excel_meta.get("client_name"):
            client_name = excel_meta["client_name"]
        if not security_tester and excel_meta.get("security_tester"):
            security_tester = excel_meta["security_tester"]
        if not reviewed_by and excel_meta.get("reviewed_by"):
            reviewed_by = excel_meta["reviewed_by"]
        if not report_date and excel_meta.get("report_date"):
            report_date = excel_meta["report_date"]

        try:
            parsed_report_date = date.fromisoformat(report_date) if report_date else date.today()
        except ValueError:
            parsed_report_date = date.today()

        metadata = EngagementMetadata(
            client_name=client_name,
            security_tester=security_tester,
            reviewed_by=reviewed_by,
            report_date=parsed_report_date,
            report_version=report_number,
            scope_label=scope,
            phase_label=phase,
            default_device_type=device_type,
            report_type=report_type,
            client_short_name=client_short_name,
            assessment_start_date=assessment_start_date,
            assessment_finish_date=assessment_finish_date,
            final_retesting_start=final_retesting_start,
            final_retesting_finish=final_retesting_finish,
            released_date=released_date,
            spokesperson_name=spokesperson_name,
            spokesperson_designation=spokesperson_designation,
            spokesperson_email=spokesperson_email,
            senior_name=senior_name,
            approved_by=approved_by,
            # New fields
            report_release_date=report_release_date,
            period=period,
            document_id=document_id,
            change_history_version=change_history_version,
            change_history_date=change_history_date,
            change_history_remarks=change_history_remarks,
            distribution_name=distribution_name,
            distribution_organization=distribution_organization,
            distribution_designation=distribution_designation,
            distribution_email=distribution_email,
            # Auditing Team - Row 1
            auditor_1_name=auditor_1_name,
            auditor_1_designation=auditor_1_designation,
            auditor_1_email=auditor_1_email,
            auditor_1_qualifications=auditor_1_qualifications,
            auditor_1_cert_in=auditor_1_cert_in,
            # Auditing Team - Row 2
            auditor_2_name=auditor_2_name,
            auditor_2_designation=auditor_2_designation,
            auditor_2_email=auditor_2_email,
            auditor_2_qualifications=auditor_2_qualifications,
            auditor_2_cert_in=auditor_2_cert_in,
        )

        project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
        word_template_path = project_root / "templates" / "Word file.docx"

        va_word_df, va_risk_summary = _read_va_data(va_path) if va_path else (pd.DataFrame(), {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Grand Total": 0})
        ca_word_df, ca_risk_summary = _read_ca_data(ca_path) if ca_path else (pd.DataFrame(), {"FAILED": 0, "WARNING": 0, "Grand Total": 0})

        scope_clean = _sanitize_filename(metadata.scope_label)
        phase_clean = _sanitize_filename(metadata.phase_label)
        client_clean = _sanitize_filename(metadata.client_name.replace(" ", "_"))
        year = str(metadata.report_date.year)
        version_str = str(metadata.report_version).replace(".", "_")
        word_parts = [
            "VA_CA",
            scope_clean,
            phase_clean,
            "Audit_Report",
            client_clean,
        ]
        if metadata.entity_codes:
            for code in metadata.entity_codes:
                word_parts.append(_sanitize_filename(code))
        word_parts.append(year)
        word_parts.append(f"V{version_str}")
        word_filename = "_".join(word_parts) + ".docx"
        output_dir = Path(tempfile.mkdtemp())
        word_output_path = output_dir / word_filename

        word_path = build_word_report(
            template_path=word_template_path,
            output_path=word_output_path,
            metadata=metadata,
            va_df=va_word_df,
            ca_df=ca_word_df,
            va_risk_summary=va_risk_summary,
            ca_risk_summary=ca_risk_summary,
        )

        session_id = create_session()
        store_file(session_id, "word", word_path)

        return JSONResponse(
            content={
                "session_id": session_id,
                "files": {"word_report": f"/api/download/{session_id}/word"},
            }
        )
    finally:
        import shutil
        for f in saved_files:
            f.unlink(missing_ok=True)
        temp_dir.rmdir()
