"""POST /api/reverse-textjoin — Convert TextJoin VA/CA report back to normal format."""

from __future__ import annotations

import logging
import sys
import tempfile
from pathlib import Path

from fastapi import APIRouter, Depends, File, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

from ..deps import get_current_user
from ..temp_registry import create_session, store_file

router = APIRouter(tags=["reverse-textjoin"])
logger = logging.getLogger("va_ca_automation")


def _detect_report_type(path: Path) -> str:
    """Detect if the file is VA or CA by checking sheet names."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        has_va = "VA Report" in wb.sheetnames
        has_ca = "CA_Report" in wb.sheetnames
        wb.close()
        if has_va:
            return "va"
        elif has_ca:
            return "ca"
    except Exception:
        pass
    return "va"


def _reverse_textjoin(input_path: Path, output_path: Path, report_type: str) -> dict:
    """Call the appropriate reverser based on report type."""
    project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
    if str(project_root) not in sys.path:
        sys.path.insert(0, str(project_root))

    if report_type == "ca":
        from CA_TextJoin_Reverser import convert
    else:
        from VA_TextJoin_Reverser import convert

    return convert(str(input_path), str(output_path), progress_cb=lambda m: logger.info(m))


@router.post("/reverse-textjoin")
async def reverse_textjoin(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """Convert a TextJoin VA/CA report back to normal per-host format."""
    temp_dir = Path(tempfile.mkdtemp())
    try:
        content = await file.read()
        suffix = Path(file.filename).suffix if file.filename else ".xlsx"
        input_path = temp_dir / (file.filename or f"upload{suffix}")
        input_path.write_bytes(content)

        report_type = _detect_report_type(input_path)

        output_dir = Path(tempfile.mkdtemp())
        base_name = Path(file.filename).stem if file.filename else "textjoin"
        output_path = output_dir / f"{base_name}_normal.xlsx"

        stats = _reverse_textjoin(input_path, output_path, report_type)

        session_id = create_session()
        file_type = "ca_normal_output" if report_type == "ca" else "va_normal_output"
        store_file(session_id, file_type, output_path)

        return JSONResponse(
            content={
                "session_id": session_id,
                "files": {"normal_report": f"/api/download/{session_id}/{file_type}"},
                "report_type": report_type,
                "stats": {
                    "old_count": stats.get("old_count", 0),
                    "new_count": stats.get("new_count", 0),
                    "risk_counts": stats.get("risk_counts", {}),
                    "status_counts": stats.get("status_counts", {}),
                },
            }
        )
    except Exception as e:
        logger.error("Reverse textjoin failed: %s", e, exc_info=True)
        return JSONResponse(status_code=500, content={"detail": str(e)})
    finally:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
