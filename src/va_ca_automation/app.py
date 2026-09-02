"""Top-level application entrypoint."""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import date
from pathlib import Path

import pandas as pd

from .metadata.engagement_metadata import EngagementMetadata, HostMetadata
from .pipelines.va_pipeline import run_va_pipeline
from .pipelines.ca_pipeline import run_ca_pipeline
from .word_writer.word_report_builder import build_word_report

# Resolve project root (src/va_ca_automation -> src -> project root)
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_TEMPLATE_PATH = _PROJECT_ROOT / "templates" / "va_report_template.xlsx"
DEFAULT_CA_TEMPLATE_PATH = _PROJECT_ROOT / "templates" / "ca_report_template.xlsx"


def setup_logging(verbose: bool = False) -> None:
    """Configure logging for the application."""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def build_metadata_from_args(args: argparse.Namespace) -> EngagementMetadata:
    """Build EngagementMetadata from parsed CLI arguments."""
    host_metadata = {}
    if args.host_metadata:
        for entry in args.host_metadata:
            parts = entry.split(":")
            if len(parts) == 3:
                ip, scan_type, device_type = parts
                host_metadata[ip.strip()] = HostMetadata(
                    ip_address=ip.strip(),
                    scan_type=scan_type.strip(),
                    device_type=device_type.strip(),
                )

    return EngagementMetadata(
        client_name=args.client_name,
        security_tester=args.tester,
        reviewed_by=args.reviewer,
        report_date=date.fromisoformat(args.report_date) if args.report_date else date.today(),
        report_version=args.version,
        scanner_name=args.scanner_name or "Nessus ",
        scanner_version=args.scanner_version or "10.11.4",
        report_owner=args.report_owner or "",
        scope_label=args.scope or "Server",
        phase_label=args.phase or "First",
        entity_codes=args.entity_codes or [],
        default_device_type=args.device_type or "",
        host_metadata=host_metadata,
    )


def main(argv: list[str] | None = None) -> int:
    """Run the application."""
    parser = argparse.ArgumentParser(
        prog="va-ca-automation",
        description="VA/CA Report Automation - Process Nessus exports into client-ready reports",
    )

    parser.add_argument(
        "raw_file",
        type=Path,
        help="Path to the raw Nessus export (.xlsx)",
    )
    parser.add_argument(
        "-t", "--template",
        type=Path,
        default=None,
        help="Path to the blank template (.xlsx). Defaults to templates/va_report_template.xlsx",
    )
    parser.add_argument(
        "-o", "--output-dir",
        type=Path,
        default=Path("output"),
        help="Output directory for the generated report (default: output/)",
    )
    parser.add_argument(
        "--client-name",
        required=True,
        help="Client legal name for the report",
    )
    parser.add_argument(
        "--tester",
        required=True,
        help="Security tester name",
    )
    parser.add_argument(
        "--reviewer",
        required=True,
        help="Reviewer name",
    )
    parser.add_argument(
        "--report-date",
        default=None,
        help="Report date in YYYY-MM-DD format (default: today)",
    )
    parser.add_argument(
        "--version",
        type=str,
        default="1.0",
        help="Report version number (default: 1.0)",
    )
    parser.add_argument(
        "--scanner-name",
        default=None,
        help="Scanner name (default: Nessus )",
    )
    parser.add_argument(
        "--scanner-version",
        default=None,
        help="Scanner version (default: 10.11.4)",
    )
    parser.add_argument(
        "--report-owner",
        default=None,
        help="Report owner name for Introduction sheet",
    )
    parser.add_argument(
        "--scope",
        default=None,
        help="Scope label (e.g., Server, Firewall)",
    )
    parser.add_argument(
        "--device-type",
        default=None,
        help="Device type for all hosts in the summary table (e.g., Server, Firewall)",
    )
    parser.add_argument(
        "--phase",
        default=None,
        help="Phase label (e.g., First, Retest)",
    )
    parser.add_argument(
        "--entity-codes",
        nargs="*",
        default=None,
        help="Entity codes (e.g., TSS SCPL)",
    )
    parser.add_argument(
        "--host-metadata",
        nargs="*",
        default=None,
        help="Per-host metadata as IP:ScanType:DeviceType (e.g., 192.168.1.1:Authenticated:Server)",
    )
    parser.add_argument(
        "--log-file",
        type=Path,
        default=None,
        help="Path to write structured log entries",
    )
    parser.add_argument(
        "--no-text-join",
        action="store_true",
        default=False,
        help="Disable text-joined report generation (by default both Normal and TextJoin reports are generated)",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose/debug logging",
    )

    args = parser.parse_args(argv)
    setup_logging(args.verbose)

    # Resolve template path
    template_path = args.template
    if template_path is None:
        template_path = DEFAULT_TEMPLATE_PATH

    # Resolve Word template path
    word_template_path = _PROJECT_ROOT / "templates" / "Word file.docx"

    if not args.raw_file.exists():
        print(f"Error: Raw file not found: {args.raw_file}", file=sys.stderr)
        return 1

    if not template_path.exists():
        print(f"Error: Template file not found: {template_path}", file=sys.stderr)
        print("Please provide a blank template using --template or place it at templates/va_report_template.xlsx", file=sys.stderr)
        return 1

    metadata = build_metadata_from_args(args)

    try:
        # Run VA pipeline
        va_output_path = run_va_pipeline(
            raw_file_path=args.raw_file,
            template_path=template_path,
            metadata=metadata,
            output_dir=args.output_dir,
            log_file=args.log_file,
            generate_text_join=not args.no_text_join,
        )
        va_tj_path = va_output_path.with_name(va_output_path.stem + "_TextJoin" + va_output_path.suffix)
        print(f"VA Normal report generated: {va_output_path}")
        print(f"VA TextJoin report generated: {va_tj_path}")

        # Run CA pipeline
        from .ingestion.raw_file_loader import load_raw_file
        raw_df = load_raw_file(args.raw_file)
        ca_template_path = DEFAULT_CA_TEMPLATE_PATH
        ca_output_path = run_ca_pipeline(
            raw_df=raw_df,
            metadata=metadata,
            ca_template_path=ca_template_path,
            output_dir=args.output_dir,
            log_file=args.log_file,
            generate_text_join=not args.no_text_join,
        )
        if ca_output_path:
            ca_tj_path = ca_output_path.with_name(ca_output_path.stem + "_TextJoin" + ca_output_path.suffix)
            print(f"CA Normal report generated: {ca_output_path}")
            print(f"CA TextJoin report generated: {ca_tj_path}")
        else:
            print("No CA findings to process.")

        # Generate Word report
        if word_template_path.exists():
            from openpyxl import load_workbook

            # Build Word filename
            from .naming.filename_builder import _sanitize_filename
            scope = _sanitize_filename(metadata.scope_label)
            phase = _sanitize_filename(metadata.phase_label)
            client_clean = _sanitize_filename(metadata.client_name.replace(" ", "_"))
            year = str(metadata.report_date.year)
            version_str = str(metadata.report_version).replace(".", "_")
            version_part = f"V{version_str}"

            word_parts = ["VA_CA", scope, phase, "Audit_Report", client_clean]
            if metadata.entity_codes:
                for code in metadata.entity_codes:
                    word_parts.append(_sanitize_filename(code))
            word_parts.append(year)
            word_parts.append(version_part)
            word_filename = "_".join(word_parts) + ".docx"
            word_output_path = args.output_dir / word_filename

            # Read VA data from generated Excel
            va_word_df = pd.DataFrame()
            va_risk_summary = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Grand Total": 0}
            try:
                va_wb = load_workbook(va_output_path, read_only=True, data_only=True)
                va_ws = va_wb["VA Report"]
                va_data = []
                for row in va_ws.iter_rows(min_row=14, max_col=9, values_only=True):
                    if row[0] is not None:
                        va_data.append(row)
                if va_data:
                    va_cols = ["Sr. no", "Vulnerbility Title", "Description", "Risk", "Host", "Port", "Recommendation ", "Reference", "CVE"]
                    va_word_df = pd.DataFrame(va_data, columns=va_cols)
                va_wb.close()

                va_summary_wb = load_workbook(va_output_path, read_only=True, data_only=True)
                va_summary_ws = va_summary_wb["Summary"]
                for row in va_summary_ws.iter_rows(min_row=16, max_row=20, min_col=5, max_col=6, values_only=True):
                    if row[0] and row[1] is not None:
                        label = str(row[0]).strip()
                        count = int(row[1]) if row[1] else 0
                        if label in va_risk_summary:
                            va_risk_summary[label] = count
                        elif label == "Grand Total":
                            va_risk_summary["Grand Total"] = count
                va_summary_wb.close()
            except Exception as e:
                logging.getLogger("va_ca_automation").warning("Could not read VA Excel for Word report: %s", e)

            # Read CA data from generated Excel
            ca_word_df = pd.DataFrame()
            ca_risk_summary = {"FAILED": 0, "WARNING": 0, "Grand Total": 0}
            if ca_output_path and ca_output_path.exists():
                try:
                    ca_wb = load_workbook(ca_output_path, read_only=True, data_only=True)
                    ca_ws = ca_wb["CA_Report"]
                    ca_data = []
                    for row in ca_ws.iter_rows(min_row=14, max_col=6, values_only=True):
                        if row[0] is not None:
                            ca_data.append(row)
                    if ca_data:
                        ca_cols = ["Sr.No.", "Title", "Host", "Description", "Solution", "Risk"]
                        ca_word_df = pd.DataFrame(ca_data, columns=ca_cols)
                    ca_wb.close()

                    ca_summary_wb = load_workbook(ca_output_path, read_only=True, data_only=True)
                    ca_summary_ws = ca_summary_wb["Summary"]
                    for row in ca_summary_ws.iter_rows(min_row=16, max_row=18, min_col=5, max_col=6, values_only=True):
                        if row[0] and row[1] is not None:
                            label = str(row[0]).strip()
                            count = int(row[1]) if row[1] else 0
                            if label in ca_risk_summary:
                                ca_risk_summary[label] = count
                            elif label == "Grand Total":
                                ca_risk_summary["Grand Total"] = count
                    ca_summary_wb.close()
                except Exception as e:
                    logging.getLogger("va_ca_automation").warning("Could not read CA Excel for Word report: %s", e)

            word_path = build_word_report(
                template_path=word_template_path,
                output_path=word_output_path,
                metadata=metadata,
                va_df=va_word_df,
                ca_df=ca_word_df,
                va_risk_summary=va_risk_summary,
                ca_risk_summary=ca_risk_summary,
            )
            print(f"Word report generated: {word_path}")
        else:
            print(f"Word template not found at {word_template_path}, skipping Word report generation.")

        return 0
    except Exception as e:
        logging.getLogger("va_ca_automation").error("Pipeline failed: %s", e, exc_info=True)
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
