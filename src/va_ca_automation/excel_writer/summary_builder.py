"""Build scope tables and summary aggregations."""

from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..metadata.engagement_metadata import EngagementMetadata

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DATA_FONT = Font(name="Cambria", size=11)

HEADER_FONT = Font(name="Cambria", size=11, bold=True)

RISK_ORDER = ["Critical", "High", "Medium", "Low"]

CA_RISK_ORDER = ["FAILED", "WARNING"]

GRAND_TOTAL_FILL = PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid")

HEADER_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")


def _split_hosts(hosts_series: pd.Series) -> list[str]:
    """Split comma-joined host strings into individual IPs and return sorted unique list."""
    seen = []
    for entry in hosts_series.dropna():
        for ip in str(entry).split(","):
            ip = ip.strip()
            if ip and ip not in seen:
                seen.append(ip)
    return sorted(seen)


def build_scope_table(va_sorted_df: pd.DataFrame, metadata: EngagementMetadata) -> pd.DataFrame:
    """Build the Summary scope table from distinct hosts in the processed data."""
    distinct_hosts = _split_hosts(va_sorted_df["Host"])

    rows = []
    for ip in distinct_hosts:
        rows.append(
            {
                "IP Address": ip,
                "Scan Type": metadata.get_host_scan_type(ip),
                "Device Type": metadata.get_host_device_type(ip),
            }
        )

    return pd.DataFrame(rows)


def write_scope_table(ws, scope_df: pd.DataFrame) -> int:
    """Write the scope table to the Summary sheet starting at row 8.

    Returns the last row written (so the caller can position the next table below it).
    """
    last_row = 7
    for i, (_, row) in enumerate(scope_df.iterrows()):
        excel_row = 8 + i
        ws.cell(row=excel_row, column=1, value=i + 1)
        ws.cell(row=excel_row, column=2, value=row["IP Address"])
        ws.cell(row=excel_row, column=3, value=row["Device Type"])
        ws.cell(row=excel_row, column=4, value=row["Scan Type"])

        for col in range(1, 5):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col != 4:
                cell.border = THIN_BORDER
        last_row = excel_row
    return last_row


def build_risk_summary(va_sorted_df: pd.DataFrame) -> dict[str, int]:
    """Build the risk count aggregation from processed VA data."""
    risk_counts = va_sorted_df["Risk"].value_counts()
    result = {}
    for risk in RISK_ORDER:
        result[risk] = int(risk_counts.get(risk, 0))
    result["Grand Total"] = sum(result.values())
    return result


def write_risk_summary_table(ws, risk_summary: dict[str, int], start_row: int = 15) -> None:
    """Write the risk summary table to the Summary sheet.

    Positions the table below the scope table with a gap, defaulting to row 15
    to match the completed-sample layout.

    Parameters
    ----------
    ws : worksheet
        The Summary worksheet.
    risk_summary : dict
        Risk level counts including "Grand Total".
    start_row : int
        The row where the header ("Row Labels" / "Count of Host") is written.
    """
    # Write header row
    ws.cell(row=start_row, column=5, value="Row Labels")
    ws.cell(row=start_row, column=6, value="Count of Host")

    # Apply border and background color to header row
    for col in [5, 6]:
        cell = ws.cell(row=start_row, column=col)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

    # Write data rows
    data_start = start_row + 1
    for i, (label, count) in enumerate(risk_summary.items()):
        ws.cell(row=data_start + i, column=5, value=label)
        ws.cell(row=data_start + i, column=6, value=count)

        for col in [5, 6]:
            cell = ws.cell(row=data_start + i, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    # Bold the Grand Total row and apply background color
    grand_total_row = data_start + len(risk_summary) - 1
    for col in [5, 6]:
        cell = ws.cell(row=grand_total_row, column=col)
        cell.font = Font(name="Cambria", size=11, bold=True)
        cell.fill = GRAND_TOTAL_FILL


# =========================================================
# CA SUMMARY FUNCTIONS
# =========================================================


def format_ca_summary_title(ws) -> None:
    """Format the CA Summary title cell (A4) with wrap_text and center alignment.

    The title 'List of Ips in scope' is in merged cells A4:C6.
    """
    cell = ws["A4"]
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )


def build_ca_scope_table(ca_df: pd.DataFrame, metadata: EngagementMetadata) -> pd.DataFrame:
    """Build the CA Summary scope table from distinct hosts in the processed data.

    Unlike the VA scope table, the CA scope table does not include a Scan Type column.
    """
    distinct_hosts = _split_hosts(ca_df["Host"])

    rows = []
    for ip in distinct_hosts:
        rows.append(
            {
                "IP Address": ip,
                "Device Type": metadata.get_host_device_type(ip),
            }
        )

    return pd.DataFrame(rows)


def write_ca_scope_table(ws, scope_df: pd.DataFrame) -> int:
    """Write the scope table to the CA Summary sheet starting at row 8.

    The CA Summary sheet has 3 columns: Sr No, IP Address, Device Type (no Scan Type).
    Returns the last row written.
    """
    last_row = 7
    for i, (_, row) in enumerate(scope_df.iterrows()):
        excel_row = 8 + i
        ws.cell(row=excel_row, column=1, value=i + 1)
        ws.cell(row=excel_row, column=2, value=row["IP Address"])
        ws.cell(row=excel_row, column=3, value=row["Device Type"])

        for col in range(1, 4):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        last_row = excel_row
    return last_row


def build_ca_risk_summary(ca_df: pd.DataFrame) -> dict[str, int]:
    """Build the risk count aggregation from processed CA data.

    CA findings use FAILED and WARNING as risk levels.
    """
    risk_counts = ca_df["Risk"].str.upper().replace("WARNINGS", "WARNING").value_counts()
    result = {}
    for risk in CA_RISK_ORDER:
        result[risk] = int(risk_counts.get(risk, 0))
    result["Grand Total"] = sum(result.values())
    return result


def write_ca_risk_summary_table(ws, risk_summary: dict[str, int], start_row: int = 15) -> None:
    """Write the risk summary table to the CA Summary sheet.

    Positions the table below the scope table with a gap, defaulting to row 15.
    """
    # Write header row
    ws.cell(row=start_row, column=5, value="Row Labels")
    ws.cell(row=start_row, column=6, value="Count of Host")

    # Apply border and background color to header row
    for col in [5, 6]:
        cell = ws.cell(row=start_row, column=col)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL

    # Write data rows
    data_start = start_row + 1
    for i, (label, count) in enumerate(risk_summary.items()):
        ws.cell(row=data_start + i, column=5, value=label)
        ws.cell(row=data_start + i, column=6, value=count)

        for col in [5, 6]:
            cell = ws.cell(row=data_start + i, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    # Bold the Grand Total row and apply background color
    grand_total_row = data_start + len(risk_summary) - 1
    for col in [5, 6]:
        cell = ws.cell(row=grand_total_row, column=col)
        cell.font = Font(name="Cambria", size=11, bold=True)
        cell.fill = GRAND_TOTAL_FILL
