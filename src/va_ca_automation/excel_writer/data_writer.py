"""Write processed findings into the report template."""

from __future__ import annotations

from copy import copy

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# Template column order (must match column_mapper.py output)
TEMPLATE_COLUMNS = [
    "Sr. no",
    "Vulnerbility Title",
    "Description",
    "Risk",
    "Host",
    "Port",
    "Recommendation ",
    "Reference",
    "CVE",
]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WRAP_COLUMNS = set()
TITLE_COLUMNS = {"Vulnerbility Title"}
CENTER_COLUMNS = {"Sr. no", "Risk", "Host", "Port", "CVE"}
LEFT_COLUMNS = {"Vulnerbility Title"}
TOP_LEFT_COLUMNS = {"Description", "Recommendation ", "Reference"}

DATA_FONT = Font(name="Cambria", size=11)
HEADER_FONT = Font(name="Cambria", size=11, bold=True)
HEADER_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")


def write_va_report_header(ws, metadata) -> None:
    """Write the VA Report header metadata block (rows 5-10, column C)."""
    ws["C5"] = metadata.client_name
    ws["C6"] = metadata.security_tester
    ws["C7"] = metadata.reviewed_by
    ws["C8"] = metadata.report_date
    ws["C9"] = metadata.report_version
    ws["C10"] = metadata.scanner_name


def style_va_headers(ws, header_row: int = 13) -> None:
    """Style the VA Report headers with centered alignment and yellow background."""
    for col in range(1, 10):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_va_data_rows(ws, df: pd.DataFrame) -> int:
    """Write processed VA data rows starting at row 14.

    Returns the last row written to.
    """
    style_va_headers(ws)
    data_start_row = 14

    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = data_start_row + i
        ws.row_dimensions[excel_row].height = 110
        for j, col_name in enumerate(TEMPLATE_COLUMNS):
            cell = ws.cell(row=excel_row, column=j + 1)

            value = row.get(col_name)
            if pd.isna(value):
                cell.value = "N/A"
            elif value == "":
                cell.value = "N/A"
            else:
                cell.value = value

            _apply_data_cell_style(cell, col_name)

    return data_start_row + len(df) - 1 if len(df) > 0 else data_start_row - 1


def _apply_data_cell_style(cell, col_name: str) -> None:
    """Apply Cambria 11pt font, thin border, and centered alignment for all columns."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER

    if col_name in LEFT_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    elif col_name in TOP_LEFT_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
    elif col_name in WRAP_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def clone_row_style_from_template(ws, source_row: int = 13) -> dict:
    """Capture the style properties from a template row for cloning.

    In a pristine template, row 13 has headers. We capture styles from
    the header-adjacent style to apply to data rows.
    """
    styles = {}
    for col in range(1, 10):
        cell = ws.cell(row=source_row, column=col)
        styles[col] = {
            "font": copy(cell.font) if cell.font else None,
            "border": copy(cell.border) if cell.border else None,
            "alignment": copy(cell.alignment) if cell.alignment else None,
        }
    return styles


def write_introduction_fields(ws, metadata) -> None:
    """Write dynamic fields to the Introduction sheet."""
    if metadata.scanner_name:
        ws["B14"] = metadata.scanner_name
    if metadata.scanner_version:
        ws["B15"] = metadata.scanner_version
    if metadata.report_owner:
        ws["B19"] = metadata.report_owner


# =========================================================
# CA REPORT WRITERS
# =========================================================

CA_TEMPLATE_COLUMNS = [
    "Sr.No.",
    "Title",
    "Host",
    "Description",
    "Solution",
    "Risk",
]

CA_WRAP_COLUMNS = {"Description", "Solution"}
CA_CENTER_COLUMNS = {"Sr.No.", "Risk"}
CA_HOST_COLUMNS = {"Host"}
CA_TITLE_COLUMNS = {"Title"}

CA_RISK_FILLS = {
    "WARNING": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
    "FAILED": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
}
CA_RISK_FONT_WHITE = Font(name="Cambria", size=11, bold=True, color="FFFFFF")


def write_ca_report_header(ws, metadata) -> None:
    """Write the CA Report header metadata block (rows 5-10, columns B-C)."""
    ws["C5"] = metadata.client_name
    ws["C6"] = metadata.security_tester
    ws["C7"] = metadata.reviewed_by
    ws["C8"] = metadata.report_date
    ws["C9"] = metadata.report_version
    ws["C10"] = metadata.scanner_name


def write_ca_data_rows(ws, df: pd.DataFrame) -> int:
    """Write processed CA data rows starting at row 14.

    The CA_Report sheet already has headers at row 13, so data starts at row 14.
    Returns the last row written to.
    """
    data_start_row = 14

    for i, (_, row) in enumerate(df.iterrows()):
        excel_row = data_start_row + i
        ws.row_dimensions[excel_row].height = 110
        for j, col_name in enumerate(CA_TEMPLATE_COLUMNS):
            cell = ws.cell(row=excel_row, column=j + 1)

            value = row.get(col_name)
            if pd.isna(value):
                cell.value = "N/A"
            elif value == "":
                cell.value = "N/A"
            else:
                cell.value = value

            _apply_ca_data_cell_style(cell, col_name)

    return data_start_row + len(df) - 1 if len(df) > 0 else data_start_row - 1


def _apply_ca_data_cell_style(cell, col_name: str) -> None:
    """Apply Cambria 11pt font, thin border, and wrap_text for CA data cells."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER

    if col_name in CA_TITLE_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    elif col_name in CA_WRAP_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    elif col_name in CA_HOST_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    elif col_name in CA_CENTER_COLUMNS:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(vertical="top")

    # Apply Risk column fill and white font
    if col_name == "Risk":
        risk_value = str(cell.value).strip().upper() if cell.value else ""
        fill = CA_RISK_FILLS.get(risk_value)
        if fill:
            cell.fill = fill
            cell.font = CA_RISK_FONT_WHITE
