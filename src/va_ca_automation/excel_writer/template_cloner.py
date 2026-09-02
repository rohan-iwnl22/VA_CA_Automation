"""Clone pristine report templates into working files."""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook


class TemplatePristencyError(Exception):
    """Raised when a template file is not in the expected pristine state."""


def clone_template(template_path: Path, output_path: Path) -> Path:
    """Copy the pristine template to a working output path.

    Verifies that the template is pristine (no data rows, no pivot tables)
    before copying.

    Returns the path to the working copy.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        shutil.copy2(template_path, output_path)
    except PermissionError:
        raise PermissionError(
            f"Cannot read template file: {template_path}\n"
            f"The file may be open in Excel or another program. "
            f"Please close it and try again."
        )

    _assert_template_is_pristine(output_path)

    return output_path


def _assert_template_is_pristine(template_path: Path) -> None:
    """Fail loudly if the template contains unexpected data rows or pivot artifacts.

    Template sheets have expected header/label rows:
    - VA Report: labels at rows 5-10, column headers at row 13 (data starts at row 14)
    - Summary: labels at rows 5 and 7 (scope table starts at row 8)
    """
    wb = load_workbook(template_path, read_only=True)

    try:
        # Check VA Report sheet has no data rows beyond header (row 13)
        if "VA Report" in wb.sheetnames:
            ws = wb["VA Report"]
            data_row_count = _count_data_rows_after(ws, after_row=13)
            if data_row_count > 0:
                raise TemplatePristencyError(
                    f"Template VA Report sheet contains {data_row_count} data rows "
                    f"beyond row 13 (expected 0 for pristine template). "
                    f"Template may contain leftover data."
                )

        # Check Summary sheet has no data rows beyond headers (row 7)
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            data_row_count = _count_data_rows_after(ws, after_row=7)
            if data_row_count > 0:
                raise TemplatePristencyError(
                    f"Template Summary sheet contains {data_row_count} data rows "
                    f"beyond row 7 (expected 0 for pristine template). "
                    f"Template may contain leftover data."
                )

        # Check no pivot tables exist (openpyxl can detect these)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if hasattr(ws, "_pivots") and ws._pivots:
                raise TemplatePristencyError(
                    f"Template sheet '{sheet_name}' contains existing pivot tables. "
                    f"Expected a pristine template with no pivot artifacts."
                )
    finally:
        wb.close()


def _count_data_rows_after(ws, after_row: int) -> int:
    """Count rows with at least one non-None cell value, starting after the given row."""
    count = 0
    for row in ws.iter_rows(min_row=after_row + 1):
        if any(cell.value is not None for cell in row):
            count += 1
    return count


def load_working_copy(working_path: Path):
    """Load the working copy workbook for editing."""
    return load_workbook(working_path)
