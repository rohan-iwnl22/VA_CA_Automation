"""Generate Consolidated Vulnerability Report (2 files).

Outputs two Excel files:
1. Normal report      - one row per host per vulnerability (no grouping)
2. TextJoin report    - IPs merged per vulnerability (TEXTJOIN-style)

Usage:
    python generate_consolidated_vuln_report.py RAW_SERVER.xlsx
    python generate_consolidated_vuln_report.py RAW_SERVER.xlsx -o output/
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

# ── Styling ──────────────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FONT = Font(name="Cambria", size=11, bold=True)
DATA_FONT = Font(name="Cambria", size=11)

WRAP_COLUMNS = {"Description", "Recommendation", "Reference"}
CENTER_COLUMNS = {"Sr. no", "Risk", "IP Address", "Port", "CVE"}

TEMPLATE_COLUMNS = [
    "Sr. no",
    "Vulnerability Name",
    "Description",
    "Risk",
    "IP Address",
    "Port",
    "Recommendation",
    "Reference",
    "CVE",
]

VA_ONLY_RISKS = {"Critical", "High", "Medium", "Low"}

EXCLUDE_NAMES = {
    "SSL Certificate Cannot Be Trusted",
    "SSL Self-Signed Certificate",
}

RISK_WEIGHTS = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

VERSION_RE = re.compile(r"(\d+(?:\.\d+){1,4})")


# ── Load & Clean ─────────────────────────────────────────────────────────
def load_and_clean(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    for col in ["Risk", "Host", "Name"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    df = df.dropna(subset=["Name", "Host"])
    df = df[df["Name"] != ""]
    df = df[df["Host"] != ""]
    return df


def filter_va_only(df: pd.DataFrame) -> pd.DataFrame:
    df = df[df["Risk"].isin(VA_ONLY_RISKS)].copy()
    df = df[~df["Name"].isin(EXCLUDE_NAMES)].copy()
    return df


def exact_dedup(df: pd.DataFrame) -> pd.DataFrame:
    keys = ["Name", "Description", "Risk", "Host"]
    return df.drop_duplicates(subset=keys, keep="first").copy()


# ── Version Collapse ─────────────────────────────────────────────────────
def _extract_version(text: str):
    m = VERSION_RE.findall(text)
    return m[-1] if m else None


def _base_title(text: str) -> str:
    return VERSION_RE.sub("", text).strip()


def _version_tuple(v: str) -> tuple:
    parts = v.split(".")
    nums = [int(p) for p in parts]
    while len(nums) < 4:
        nums.append(0)
    return tuple(nums)


def version_collapse(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_ver"] = df["Name"].apply(_extract_version)
    df["_base"] = df["Name"].apply(_base_title)

    kept = []
    for (_, risk, host), group in df.groupby(["_base", "Risk", "Host"]):
        no_ver = group[group["_ver"].isna()]
        ver = group[group["_ver"].notna()]
        if not no_ver.empty:
            kept.append(no_ver)
        if ver.empty:
            continue
        if len(ver) == 1:
            kept.append(ver)
            continue
        ver = ver.copy()
        ver["_vtup"] = ver["_ver"].apply(_version_tuple)
        ver_sorted = ver.sort_values("_vtup", ascending=False)
        kept.append(ver_sorted.iloc[[0]])

    result = (
        pd.concat(kept, ignore_index=True) if kept else pd.DataFrame(columns=df.columns)
    )
    return result.drop(columns=["_ver", "_base"], errors="ignore")


# ── Column Mapping ───────────────────────────────────────────────────────
RAW_MAP = {
    "Name": "Vulnerability Name",
    "Description": "Description",
    "Risk": "Risk",
    "Host": "IP Address",
    "Port": "Port",
    "Solution": "Recommendation",
    "See Also": "Reference",
    "CVE": "CVE",
}


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapped = pd.DataFrame()
    for raw_col, tpl_col in RAW_MAP.items():
        mapped[tpl_col] = df[raw_col] if raw_col in df.columns else ""
    for col in ["Reference", "CVE"]:
        mapped[col] = mapped[col].fillna("N/A").replace("", "N/A")
    mapped["Port"] = pd.to_numeric(mapped["Port"], errors="coerce").fillna(0).astype(int)
    return mapped


# ── Consolidation: group by vulnerability, join IPs ─────────────────────
def consolidate_vulns(df: pd.DataFrame) -> pd.DataFrame:
    """Group rows by Vulnerability Name and join unique IP Addresses.

    For rows sharing the same Vulnerability Name:
    - Joins all unique IP Address values with ", " into a single string
    - Takes the first non-null value for all other columns
    - Returns one row per unique vulnerability
    """
    join_cols = ["Description", "Risk", "Port", "Recommendation", "Reference", "CVE"]

    def first_non_null(s):
        non_null = s.dropna()
        return non_null.iloc[0] if not non_null.empty else ""

    def unique_ips(ips):
        seen = []
        for ip in ips.dropna():
            if ip not in seen:
                seen.append(ip)
        return ", ".join(seen) if seen else ""

    agg = {}
    for col in join_cols:
        if col in df.columns:
            agg[col] = first_non_null
    if "IP Address" in df.columns:
        agg["IP Address"] = unique_ips

    grouped = df.groupby("Vulnerability Name", sort=False).agg(agg).reset_index()

    grouped["_rw"] = grouped["Risk"].map(RISK_WEIGHTS)
    grouped = grouped.sort_values(by=["_rw"], kind="stable").drop(columns=["_rw"])
    grouped = grouped.reset_index(drop=True)
    grouped.insert(0, "Sr. no", range(1, len(grouped) + 1))
    return grouped


# ── Host Count Summary ───────────────────────────────────────────────────
def _write_host_summary(ws, data_df: pd.DataFrame, start_row: int = 15):
    """Write host count summary table at specified row in columns E-F."""
    risk_order = ["Critical", "High", "Medium", "Low"]
    host_counts = {}
    for risk in risk_order:
        risk_df = data_df[data_df["Risk"] == risk]
        host_counts[risk] = risk_df["IP Address"].nunique()
    grand_total = data_df["IP Address"].nunique()

    summary_data = [("Row Labels", "Count of Host")] + [(r, host_counts[r]) for r in risk_order] + [("Grand Total", grand_total)]

    for i, (label, count) in enumerate(summary_data):
        row_num = start_row + i
        cell_label = ws.cell(row=row_num, column=5, value=label)
        cell_count = ws.cell(row=row_num, column=6, value=count)
        cell_label.font = HEADER_FONT if i == 0 else DATA_FONT
        cell_count.font = HEADER_FONT if i == 0 else DATA_FONT
        cell_label.border = THIN_BORDER
        cell_count.border = THIN_BORDER
        cell_label.alignment = Alignment(horizontal="center", vertical="center")
        cell_count.alignment = Alignment(horizontal="center", vertical="center")
        if i == len(summary_data) - 1:
            cell_label.font = Font(name="Cambria", size=11, bold=True)
            cell_count.font = Font(name="Cambria", size=11, bold=True)


# ── Write to Excel ───────────────────────────────────────────────────────
def _apply_style(cell, col_name: str, is_textjoin: bool = False):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if col_name == "IP Address" and is_textjoin:
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
    elif col_name in WRAP_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    elif col_name in CENTER_COLUMNS:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(vertical="top")


def write_normal_report(data_df: pd.DataFrame, output_path: Path) -> Path:
    """Write normal report - one row per host per vulnerability."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (_, row) in enumerate(data_df.iterrows()):
        excel_row = i + 2
        for j, col_name in enumerate(TEMPLATE_COLUMNS):
            cell = ws.cell(row=excel_row, column=j + 1)
            value = row.get(col_name)
            if pd.isna(value) or value == "" or value is None:
                cell.value = "N/A"
            else:
                cell.value = value
            _apply_style(cell, col_name)

    ws.column_dimensions["A"].width = 8    # Sr. no
    ws.column_dimensions["B"].width = 45   # Vulnerability Name
    ws.column_dimensions["C"].width = 40   # Description
    ws.column_dimensions["D"].width = 12   # Risk
    ws.column_dimensions["E"].width = 55   # IP Address
    ws.column_dimensions["F"].width = 10   # Port
    ws.column_dimensions["G"].width = 40   # Recommendation
    ws.column_dimensions["H"].width = 30   # Reference
    ws.column_dimensions["I"].width = 25   # CVE

    _write_host_summary(ws, data_df, start_row=15)

    wb.save(output_path)
    wb.close()
    return output_path


def write_textjoin_report(data_df: pd.DataFrame, output_path: Path) -> Path:
    """Write TextJoin report - IPs consolidated per vulnerability."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated Vulnerabilities"

    for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (_, row) in enumerate(data_df.iterrows()):
        excel_row = i + 2
        for j, col_name in enumerate(TEMPLATE_COLUMNS):
            cell = ws.cell(row=excel_row, column=j + 1)
            value = row.get(col_name)
            if pd.isna(value) or value == "" or value is None:
                cell.value = "N/A"
            else:
                cell.value = value
            _apply_style(cell, col_name, is_textjoin=True)

    ws.column_dimensions["A"].width = 8    # Sr. no
    ws.column_dimensions["B"].width = 45   # Vulnerability Name
    ws.column_dimensions["C"].width = 40   # Description
    ws.column_dimensions["D"].width = 12   # Risk
    ws.column_dimensions["E"].width = 55   # IP Address
    ws.column_dimensions["F"].width = 10   # Port
    ws.column_dimensions["G"].width = 40   # Recommendation
    ws.column_dimensions["H"].width = 30   # Reference
    ws.column_dimensions["I"].width = 25   # CVE

    _write_host_summary(ws, data_df, start_row=15)

    wb.save(output_path)
    wb.close()
    return output_path


# ── Main ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Generate consolidated vulnerability report (IPs merged per vulnerability)"
    )
    parser.add_argument("raw_file", type=Path, help="Path to raw Nessus export (.xlsx)")
    parser.add_argument(
        "-o", "--output-dir", type=Path, default=Path("output"),
        help="Output directory (default: output/)",
    )
    parser.add_argument(
        "--output-name", type=str, default=None,
        help="Output filename (without extension). Defaults to raw file name.",
    )
    args = parser.parse_args()

    if not args.raw_file.exists():
        print(f"Error: Raw file not found: {args.raw_file}", file=sys.stderr)
        return 1

    args.output_dir.mkdir(parents=True, exist_ok=True)

    base_name = args.output_name or "VA_Server_First_Audit_Report_SCPL_2026_V1_0"
    if base_name.endswith(".xlsx"):
        base_name = base_name[:-5]
    normal_path = args.output_dir / f"{base_name}_Normal.xlsx"
    textjoin_path = args.output_dir / f"{base_name}_TextJoin.xlsx"

    print("Loading raw data...")
    raw = load_and_clean(args.raw_file)
    print(f"  Raw rows: {len(raw)}")

    print("Filtering to VA risks only...")
    va_only = filter_va_only(raw)
    print(f"  After filter: {len(va_only)}")

    print("Exact dedup...")
    s1 = exact_dedup(va_only)
    print(f"  After dedup: {len(s1)}")

    print("Name+Host dedup...")
    s1b = s1.drop_duplicates(subset=["Name", "Host"], keep="first").copy()
    print(f"  After Name+Host dedup: {len(s1b)}")

    print("Version collapse...")
    s2 = version_collapse(s1b)
    print(f"  After version collapse: {len(s2)}")

    print("Mapping columns...")
    mapped = map_columns(s2)

    print(f"Writing normal report: {normal_path.name}")
    write_normal_report(mapped, normal_path)
    print(f"  Saved: {normal_path}")

    print("Consolidating vulnerabilities (TEXTJOIN-style)...")
    consolidated = consolidate_vulns(mapped)
    print(f"  Unique vulnerabilities: {len(consolidated)}")
    print(f"  Risk breakdown: {consolidated['Risk'].value_counts().to_dict()}")

    print(f"Writing TextJoin report: {textjoin_path.name}")
    write_textjoin_report(consolidated, textjoin_path)
    print(f"  Saved: {textjoin_path}")

    print("\nDone! 2 reports saved in output/")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
