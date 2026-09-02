"""Generate TextJoin VA Report - Groups IPs by Vulnerability Name.

This script processes a raw Nessus export and generates an Excel report
where rows sharing the same Vulnerability Title are collapsed, with all
unique Host IPs joined as a comma-separated string.

Only Critical, High, Medium, and Low risks are included.

Usage:
    python generate_textjoin_report.py RAW_SERVER.xlsx
    python generate_textjoin_report.py RAW_SERVER.xlsx -o output/ -t templates/va_report_template.xlsx
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from src.va_ca_automation.excel_writer.chart_builder import build_pie_chart
from src.va_ca_automation.excel_writer.data_writer import style_va_headers

# ── Styling constants ───────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
DATA_FONT = Font(name="Cambria", size=11)
WRAP_COLUMNS = {"Description", "Recommendation ", "Reference"}
TITLE_COLUMNS = {"Vulnerbility Title"}
CENTER_COLUMNS = {"Sr. no", "Risk", "Host", "Port", "CVE"}

TEMPLATE_COLUMNS = [
    "Sr. no",
    "Vulnerbility Title",
    "Description",
    "Risk",
    "Host",
    "Port",
    "Recommendation ",
    "Reference",
    "CVE",
]

VA_ONLY_RISKS = {"Critical", "High", "Medium", "Low"}

EXCLUDE_NAMES = {
    "SSL Certificate Cannot Be Trusted",
    "SSL Self-Signed Certificate",
}

RISK_WEIGHTS = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

VERSION_RE = re.compile(r"(\d+(?:\.\d+){1,4})")


# ── Data Processing ─────────────────────────────────────────────────────
def load_and_clean(path: Path) -> pd.DataFrame:
    """Load raw Nessus export and clean whitespace."""
    df = pd.read_excel(path)
    for col in ["Risk", "Host", "Name"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    df = df.dropna(subset=["Name", "Host"])
    df = df[df["Name"] != ""]
    df = df[df["Host"] != ""]
    return df


def filter_va_only(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only Critical, High, Medium, Low risk rows and exclude informational SSL findings."""
    df = df[df["Risk"].isin(VA_ONLY_RISKS)].copy()
    df = df[~df["Name"].isin(EXCLUDE_NAMES)].copy()
    return df


def exact_dedup(df: pd.DataFrame) -> pd.DataFrame:
    """Remove exact duplicate rows based on Name, Description, Risk, Host."""
    keys = ["Name", "Description", "Risk", "Host"]
    return df.drop_duplicates(subset=keys, keep="first").copy()


def _extract_version(text: str):
    """Extract version number from vulnerability name."""
    m = VERSION_RE.findall(text)
    return m[-1] if m else None


def _base_title(text: str) -> str:
    """Remove version numbers from vulnerability name."""
    return VERSION_RE.sub("", text).strip()


def _version_tuple(v: str) -> tuple:
    """Convert version string to comparable tuple."""
    parts = v.split(".")
    nums = [int(p) for p in parts]
    while len(nums) < 4:
        nums.append(0)
    return tuple(nums)


def version_collapse(df: pd.DataFrame) -> pd.DataFrame:
    """Collapse multiple version variants of the same vulnerability."""
    df = df.copy()
    df["_ver"] = df["Name"].apply(_extract_version)
    df["_base"] = df["Name"].apply(_base_title)

    kept = []
    for (_, risk, host), group in df.groupby(["_base", "Risk", "Host"]):
        no_ver = group[group["_ver"].isna()]
        ver = group[group["_ver"].notna()]
        if not no_ver.empty:
            kept.append(no_ver)
        if ver.empty:
            continue
        if len(ver) == 1:
            kept.append(ver)
            continue
        ver = ver.copy()
        ver["_vtup"] = ver["_ver"].apply(_version_tuple)
        ver_sorted = ver.sort_values("_vtup", ascending=False)
        kept.append(ver_sorted.iloc[[0]])

    result = (
        pd.concat(kept, ignore_index=True) if kept else pd.DataFrame(columns=df.columns)
    )
    return result.drop(columns=["_ver", "_base"], errors="ignore")


# ── Column Mapping ──────────────────────────────────────────────────────
RAW_MAP = {
    "Name": "Vulnerbility Title",
    "Description": "Description",
    "Risk": "Risk",
    "Host": "Host",
    "Port": "Port",
    "Solution": "Recommendation ",
    "See Also": "Reference",
    "CVE": "CVE",
}


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Map raw Nessus columns to template column names."""
    mapped = pd.DataFrame()
    for raw_col, tpl_col in RAW_MAP.items():
        mapped[tpl_col] = df[raw_col] if raw_col in df.columns else ""
    for col in ["Reference", "CVE"]:
        mapped[col] = mapped[col].fillna("N/A").replace("", "N/A")
    mapped["Port"] = pd.to_numeric(mapped["Port"], errors="coerce").fillna(0).astype(int)
    return mapped


# ── Sorting ─────────────────────────────────────────────────────────────
def sort_data(df: pd.DataFrame) -> pd.DataFrame:
    """Sort by host grouping then risk severity."""
    df = df.copy()
    host_order = {}
    c = 0
    for h in df["Host"]:
        if h not in host_order:
            host_order[h] = c
            c += 1
    df["_ho"] = df["Host"].map(host_order)
    df["_rw"] = df["Risk"].map(RISK_WEIGHTS)
    df = df.sort_values(by=["_ho", "_rw"], kind="stable").drop(columns=["_ho", "_rw"])
    df = df.reset_index(drop=True)
    df.insert(0, "Sr. no", range(1, len(df) + 1))
    return df


# ── TextJoin: group by vulnerability, join IPs ──────────────────────────
def text_join(df: pd.DataFrame) -> pd.DataFrame:
    """Group rows by Vulnerbility Title and join unique Host IPs with comma separator.

    This is the core function that implements the TextJoin logic:
    - Groups all rows with the same Vulnerability Title
    - Joins all unique Host IPs into a comma-separated string
    - Takes the first non-null value for other columns
    - Returns one row per unique vulnerability with merged IPs
    """
    join_cols = ["Description", "Risk", "Port", "Recommendation ", "Reference", "CVE"]

    def first_non_null(s):
        non_null = s.dropna()
        return non_null.iloc[0] if not non_null.empty else ""

    def unique_hosts(hosts):
        seen = []
        for h in hosts.dropna():
            if h not in seen:
                seen.append(h)
        return ", ".join(seen) if seen else ""

    agg = {}
    for col in join_cols:
        if col in df.columns:
            agg[col] = first_non_null
    if "Host" in df.columns:
        agg["Host"] = unique_hosts

    grouped = df.groupby("Vulnerbility Title", sort=False).agg(agg).reset_index()

    grouped["_rw"] = grouped["Risk"].map(RISK_WEIGHTS)
    grouped = grouped.sort_values(by=["_rw"], kind="stable").drop(columns=["_rw"])
    grouped = grouped.reset_index(drop=True)
    grouped.insert(0, "Sr. no", range(1, len(grouped) + 1))
    return grouped


# ── Excel Writing ───────────────────────────────────────────────────────
def _apply_style(cell, col_name: str, is_textjoin: bool = False):
    """Apply styling to a cell based on column type."""
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if col_name in TITLE_COLUMNS:
        cell.alignment = Alignment(
            wrap_text=True, horizontal="left", vertical="center"
        )
    elif col_name in WRAP_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    elif col_name == "Host" and is_textjoin:
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
    elif col_name in CENTER_COLUMNS:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(vertical="top")


def _write_data_rows(ws, data_df: pd.DataFrame, is_textjoin: bool = False) -> None:
    """Write data rows into a worksheet starting at row 14."""
    data_start = 14
    for i, (_, row) in enumerate(data_df.iterrows()):
        excel_row = data_start + i
        for j, col_name in enumerate(TEMPLATE_COLUMNS):
            cell = ws.cell(row=excel_row, column=j + 1)
            value = row.get(col_name)
            if pd.isna(value) or value == "" or value is None:
                cell.value = "N/A"
            else:
                cell.value = value
            _apply_style(cell, col_name, is_textjoin=is_textjoin)


def _write_scope(ws, df):
    """Write unique hosts into the Summary scope table."""
    seen = []
    for entry in df["Host"].dropna():
        for ip in str(entry).split(","):
            ip = ip.strip()
            if ip and ip not in seen:
                seen.append(ip)
    hosts = sorted(seen)
    start_row = 8
    last_row = 7
    for i, host in enumerate(hosts):
        excel_row = start_row + i
        ws.cell(row=excel_row, column=1, value=i + 1)
        ws.cell(row=excel_row, column=2, value=host)
        ws.cell(row=excel_row, column=3, value="N/A")
        ws.cell(row=excel_row, column=4, value="N/A")
        for col in range(1, 5):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        last_row = excel_row
    return last_row


def _write_risk_summary(ws, df, start_row: int = 15):
    """Write risk breakdown counts into Summary."""
    counts = df["Risk"].value_counts()
    # Header
    ws.cell(row=start_row, column=5, value="Row Labels")
    ws.cell(row=start_row, column=6, value="Count of Host")
    for col in [5, 6]:
        cell = ws.cell(row=start_row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER

    row = start_row + 1
    for risk in ["Critical", "High", "Medium", "Low"]:
        count = int(counts.get(risk, 0))
        ws.cell(row=row, column=5, value=risk)
        ws.cell(row=row, column=6, value=count)
        for col in [5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        row += 1

    ws.cell(row=row, column=5, value="Grand Total")
    ws.cell(row=row, column=6, value=len(df))
    for col in [5, 6]:
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Cambria", size=11, bold=True)
        cell.border = THIN_BORDER


def write_report(template_path: Path, output_path: Path, data_df: pd.DataFrame) -> Path:
    """Clone template, write data rows, and save the report."""
    wb = load_workbook(template_path)
    ws = wb["VA Report"]
    style_va_headers(ws)
    _write_data_rows(ws, data_df, is_textjoin=True)

    summary_ws = wb["Summary"]
    _write_scope(summary_ws, data_df)
    _write_risk_summary(summary_ws, data_df, start_row=15)
    build_pie_chart(summary_ws, data_df["Risk"].value_counts().to_dict(), chart_anchor="I8", data_start_row=15)

    wb.save(output_path)
    wb.close()
    return output_path


# ── Main ────────────────────────────────────────────────────────────────
def main():
    """Main entry point for generating TextJoin report."""
    parser = argparse.ArgumentParser(
        description="Generate VA TextJoin report from Nessus export (groups IPs by vulnerability)"
    )
    parser.add_argument("raw_file", type=Path, help="Path to raw Nessus export (.xlsx)")
    parser.add_argument(
        "-t",
        "--template",
        type=Path,
        default=None,
        help="Path to blank template (.xlsx)",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=Path("output"),
        help="Output directory (default: output/)",
    )
    parser.add_argument(
        "--output-name",
        type=str,
        default=None,
        help="Base output filename (without extension). Defaults to raw file name.",
    )
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent
    template = args.template or project_root / "templates" / "va_report_template.xlsx"

    if not args.raw_file.exists():
        print(f"Error: Raw file not found: {args.raw_file}", file=sys.stderr)
        return 1
    if not template.exists():
        print(f"Error: Template not found: {template}", file=sys.stderr)
        return 1

    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Determine base filename
    if args.output_name:
        base_name = args.output_name
        if base_name.endswith(".xlsx"):
            base_name = base_name[:-5]
    else:
        base_name = args.raw_file.stem

    textjoin_path = args.output_dir / "VA_Server_First_Audit_Report_SCPL_2026_V1_0_TextJoin.xlsx"

    print("Loading raw data...")
    raw = load_and_clean(args.raw_file)
    print(f"  Raw rows: {len(raw)}")

    print("Filtering to VA risks only (Critical, High, Medium, Low)...")
    va_only = filter_va_only(raw)
    print(f"  After filter: {len(va_only)}")

    print("Stage 1: Exact dedup...")
    s1 = exact_dedup(va_only)
    print(f"  After dedup: {len(s1)}")

    print("Stage 1b: Name+Host dedup...")
    s1b = s1.drop_duplicates(subset=["Name", "Host"], keep="first").copy()
    print(f"  After Name+Host dedup: {len(s1b)}")

    print("Stage 2: Version collapse...")
    s2 = version_collapse(s1b)
    print(f"  After version collapse: {len(s2)}")

    print("Mapping columns...")
    mapped = map_columns(s2)

    print("Sorting...")
    sorted_df = sort_data(mapped)
    print(f"  Final rows: {len(sorted_df)}")
    print(f"  Unique hosts: {sorted_df['Host'].nunique()}")
    print(f"  Risk breakdown: {sorted_df['Risk'].value_counts().to_dict()}")

    # ── TextJoin report ──────────────────────────────────────────────────
    tj_df = text_join(sorted_df)
    print(f"\nGenerating TextJoin report: {textjoin_path.name}")
    print(f"  TextJoin rows: {len(tj_df)} (from {len(sorted_df)} normal rows)")
    write_report(template, textjoin_path, tj_df)
    print(f"  Saved: {textjoin_path}")

    print("\nDone!")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
