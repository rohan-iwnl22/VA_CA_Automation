"""Generate a normal Excel report from RAW_SERVER.xlsx:
  One row per host per vulnerability (no grouping/TextJoin).
"""

from pathlib import Path
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from src.va_ca_automation.excel_writer.chart_builder import build_pie_chart
from src.va_ca_automation.excel_writer.data_writer import style_va_headers

# ── Paths ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
RAW_FILE = ROOT / "RAW_SERVER.xlsx"
TEMPLATE = ROOT / "templates" / "va_report_template.xlsx"
OUTPUT_DIR = ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Styling constants (matching data_writer.py) ──────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
DATA_FONT = Font(name="Cambria", size=11)
WRAP_COLUMNS = {"Description", "Recommendation ", "Reference"}
TITLE_COLUMNS = {"Vulnerbility Title"}
CENTER_COLUMNS = {"Sr. no", "Risk", "Host", "Port", "CVE"}

TEMPLATE_COLUMNS = [
    "Sr. no", "Vulnerbility Title", "Description", "Risk",
    "Host", "Port", "Recommendation ", "Reference", "CVE",
]

# Only VA risks considered; exclude FAILED, WARNING, PASSED, None
VA_ONLY_RISKS = {"Critical", "High", "Medium", "Low"}

RISK_WEIGHTS = {
    "Critical": 0, "High": 1, "Medium": 2, "Low": 3,
}


# ── Step 1: Load and clean raw data ─────────────────────────────────────
def load_and_clean(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    # Strip whitespace from key columns
    for col in ["Risk", "Host", "Name"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    # Drop rows with no Name or Host
    df = df.dropna(subset=["Name", "Host"])
    df = df[df["Name"] != ""]
    df = df[df["Host"] != ""]
    return df


# ── Step 2: Exact dedup on (Name, Description, Risk, Host) ──────────────
def exact_dedup(df: pd.DataFrame) -> pd.DataFrame:
    keys = ["Name", "Description", "Risk", "Host"]
    return df.drop_duplicates(subset=keys, keep="first").copy()


# ── Step 3: Version-collapse dedup ──────────────────────────────────────
import re
VERSION_RE = re.compile(r'(\d+(?:\.\d+){1,4})')
VERSION_FULL_RE = re.compile(r'(\d+(?:\.\d+){1,4}[a-zA-Z_]*)')


def _extract_version(text: str):
    m = re.search(r'RHSA-(\d+):(\d+)', text)
    if m:
        return f"{m.group(1)}.{m.group(2)}"
    m = re.search(r'<\s*(\d+(?:\.\d+){1,4}[a-zA-Z_]*)', text)
    if m:
        return m.group(1)
    m = re.search(r'(?:January|February|March|April|May|June|July|August|September|October|November|December)\s*(\d{4})\s*CPU', text)
    if m:
        return f"0.{m.group(1)}"
    m = re.search(r'(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-zA-Z]*(\d{4})\s*CPU', text)
    if m:
        return f"0.{m.group(1)}"
    m = VERSION_FULL_RE.findall(text)
    return m[-1] if m else None


def _base_title(text: str) -> str:
    text = re.sub(r'Oracle Java SE\s+.*?(?:Multiple\s+Vulnerabilit\w*|Information\s+Disclosure)', 'Oracle Java SE Vulnerability', text)
    text = VERSION_FULL_RE.sub("", text)
    text = re.sub(r'RHSA-\d+:\d+', 'RHSA-XXXX:XXXX', text)
    text = re.sub(r'Multiple\s+Vulnerabilit\w*', 'Vulnerability', text)
    text = re.sub(r'\(?\s*(?:Unix\s*\(\s*)?(?:January|February|March|April|May|June|July|August|September|October|November|December)\s*\d{0,4}\s*CPU\s*\)?\s*(?:\(Unix\))?', '', text)
    text = re.sub(r'\(?\s*(?:Unix\s*\(\s*)?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-zA-Z]*\s*\d{0,4}\s*CPU\s*\)?\s*(?:\(Unix\))?', '', text)
    text = re.sub(r'\(Unix\)', '', text)
    text = re.sub(r'CVE-\d+-\d+', '', text)
    return re.sub(r'\s+', ' ', text).strip()


def _version_tuple(v: str) -> tuple:
    parts = re.split(r'[._]', v)
    result = []
    for p in parts:
        m = re.match(r'(\d+)(.*)', p)
        if m:
            result.append(int(m.group(1)))
            suffix = m.group(2)
            if suffix:
                for ch in suffix:
                    result.append(ord(ch))
        else:
            result.append(0)
    while len(result) < 8:
        result.append(0)
    return tuple(result[:8])


def version_collapse(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_ver"] = df["Name"].apply(_extract_version)
    df["_base"] = df["Name"].apply(_base_title)

    kept = []
    for (_, host), group in df.groupby(["_base", "Host"]):
        no_ver = group[group["_ver"].isna()]
        ver = group[group["_ver"].notna()]
        if not no_ver.empty:
            kept.append(no_ver)
        if ver.empty:
            continue
        if len(ver) == 1:
            kept.append(ver)
            continue
        ver = ver.copy()
        ver["_vtup"] = ver["_ver"].apply(_version_tuple)
        ver_sorted = ver.sort_values("_vtup", ascending=False)
        kept.append(ver_sorted.iloc[[0]])

    result = pd.concat(kept, ignore_index=True) if kept else pd.DataFrame(columns=df.columns)
    return result.drop(columns=["_ver", "_base"], errors="ignore")


# ── Step 4: Map raw columns → template columns ──────────────────────────
RAW_MAP = {
    "Name": "Vulnerbility Title",
    "Description": "Description",
    "Risk": "Risk",
    "Host": "Host",
    "Port": "Port",
    "Solution": "Recommendation ",
    "See Also": "Reference",
    "CVE": "CVE",
}


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapped = pd.DataFrame()
    for raw_col, tpl_col in RAW_MAP.items():
        mapped[tpl_col] = df[raw_col] if raw_col in df.columns else ""
    # Fill N/A for Reference and CVE
    for col in ["Reference", "CVE"]:
        mapped[col] = mapped[col].fillna("N/A").replace("", "N/A")
    # Port: keep as-is (empty string for missing)
    mapped["Port"] = mapped["Port"].fillna("").replace("nan", "")
    return mapped


# ── Step 5: Sort by host grouping then risk severity ────────────────────
def sort_data(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    sorted_hosts = sorted(df["Host"].unique())
    host_order = {h: i for i, h in enumerate(sorted_hosts)}
    df["_ho"] = df["Host"].map(host_order)
    df["_rw"] = df["Risk"].map(RISK_WEIGHTS)
    df = df.sort_values(by=["_ho", "_rw"], kind="stable").drop(columns=["_ho", "_rw"])
    df = df.reset_index(drop=True)
    df.insert(0, "Sr. no", range(1, len(df) + 1))
    return df


# ── Step 6: Write data into template ────────────────────────────────────
def _apply_style(cell, col_name: str):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if col_name in WRAP_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_report(template_path: Path, output_path: Path, data_df: pd.DataFrame) -> Path:
    """Clone template, write data rows, save."""
    wb = load_workbook(template_path)
    ws = wb["VA Report"]
    _write_data_rows(ws, data_df)

    # Summary sheet - scope table
    summary_ws = wb["Summary"]
    _write_scope(summary_ws, data_df)
    _write_risk_summary(summary_ws, data_df, start_row=15)

    wb.save(output_path)
    wb.close()
    return output_path


def _write_data_rows(ws, data_df: pd.DataFrame) -> None:
    """Write data rows into a worksheet starting at row 14."""
    data_start = 14
    for i, (_, row) in enumerate(data_df.iterrows()):
        excel_row = data_start + i
        ws.row_dimensions[excel_row].height = 110
        for j, col_name in enumerate(TEMPLATE_COLUMNS):
            cell = ws.cell(row=excel_row, column=j + 1)
            value = row.get(col_name)
            if col_name == "Port":
                cell.value = value if value not in (None, "nan", "") else ""
            elif pd.isna(value) or value == "" or value is None:
                cell.value = "N/A"
            else:
                cell.value = value
            _apply_style(cell, col_name)


def _write_scope(ws, df):
    """Write unique hosts into the Summary scope table (starts row 8)."""
    seen = []
    for entry in df["Host"].dropna():
        for ip in str(entry).split(","):
            ip = ip.strip()
            if ip and ip not in seen:
                seen.append(ip)
    hosts = sorted(seen)
    start_row = 8
    last_row = 7
    for i, host in enumerate(hosts):
        excel_row = start_row + i
        ws.cell(row=excel_row, column=1, value=i + 1)
        ws.cell(row=excel_row, column=2, value=host)
        ws.cell(row=excel_row, column=3, value="N/A")
        ws.cell(row=excel_row, column=4, value="N/A")
        for col in range(1, 5):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        last_row = excel_row
    return last_row


def _write_risk_summary(ws, df, start_row: int = 15):
    """Write risk breakdown counts into Summary (starts at start_row, column E)."""
    counts = df["Risk"].value_counts()
    # Header
    ws.cell(row=start_row, column=5, value="Row Labels")
    ws.cell(row=start_row, column=6, value="Count of Host")
    for col in [5, 6]:
        cell = ws.cell(row=start_row, column=col)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER

    row = start_row + 1
    for risk in ["Critical", "High", "Medium", "Low"]:
        count = int(counts.get(risk, 0))
        ws.cell(row=row, column=5, value=risk)
        ws.cell(row=row, column=6, value=count)
        for col in [5, 6]:
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
        row += 1

    # Grand Total
    ws.cell(row=row, column=5, value="Grand Total")
    ws.cell(row=row, column=6, value=len(df))
    for col in [5, 6]:
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Cambria", size=11, bold=True)
        cell.border = THIN_BORDER


# ══════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════
EXCLUDE_NAMES = {
    "SSL Certificate Cannot Be Trusted",
    "SSL Self-Signed Certificate",
}


def filter_va_only(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only Critical, High, Medium, Low risk rows and exclude informational entries."""
    df = df[df["Risk"].isin(VA_ONLY_RISKS)].copy()
    df = df[~df["Name"].isin(EXCLUDE_NAMES)].copy()
    return df


def text_join(df: pd.DataFrame) -> pd.DataFrame:
    """Group rows by Vulnerbility Title and join unique Host IPs with comma separator."""
    join_cols = ["Description", "Risk", "Port", "Recommendation ", "Reference", "CVE"]

    def first_non_null(s):
        non_null = s.dropna()
        return non_null.iloc[0] if not non_null.empty else ""

    def unique_hosts(hosts):
        seen = []
        for h in hosts.dropna():
            if h not in seen:
                seen.append(h)
        return ", ".join(seen) if seen else ""

    agg = {}
    for col in join_cols:
        if col in df.columns:
            agg[col] = first_non_null
    if "Host" in df.columns:
        agg["Host"] = unique_hosts

    grouped = df.groupby("Vulnerbility Title", sort=False).agg(agg).reset_index()

    grouped["_rw"] = grouped["Risk"].map(RISK_WEIGHTS)
    grouped = grouped.sort_values(by=["_rw"], kind="stable").drop(columns=["_rw"])
    grouped = grouped.reset_index(drop=True)
    grouped.insert(0, "Sr. no", range(1, len(grouped) + 1))
    return grouped


def main():
    print("Loading raw data...")
    raw = load_and_clean(RAW_FILE)
    print(f"  Raw rows: {len(raw)}")

    print("Filtering to VA risks only (Critical, High, Medium, Low)...")
    va_only = filter_va_only(raw)
    print(f"  After filter: {len(va_only)}")

    print("Stage 1: Exact dedup...")
    s1 = exact_dedup(va_only)
    print(f"  After dedup: {len(s1)}")

    print("Stage 1b: Name+Host dedup...")
    s1b = s1.drop_duplicates(subset=["Name", "Host"], keep="first").copy()
    print(f"  After Name+Host dedup: {len(s1b)}")

    print("Stage 2: Version collapse...")
    s2 = version_collapse(s1b)
    print(f"  After version collapse: {len(s2)}")

    print("Mapping columns...")
    mapped = map_columns(s2)

    print("Sorting...")
    sorted_df = sort_data(mapped)
    print(f"  Final rows: {len(sorted_df)}")
    print(f"  Unique hosts: {sorted_df['Host'].nunique()}")
    print(f"  Risk breakdown: {sorted_df['Risk'].value_counts().to_dict()}")

    # ── Output 1: Normal report ──────────────────────────────────────────
    normal_path = OUTPUT_DIR / "VA_Server_First_Audit_Report_SCPL_2026_V1_0.xlsx"
    print(f"\nGenerating normal report: {normal_path.name}")

    wb = load_workbook(TEMPLATE)
    ws_normal = wb["VA Report"]
    style_va_headers(ws_normal)
    _write_data_rows(ws_normal, sorted_df)
    print(f"  VA Report sheet: {len(sorted_df)} rows")

    summary_ws = wb["Summary"]
    _write_scope(summary_ws, sorted_df)
    _write_risk_summary(summary_ws, sorted_df, start_row=15)
    build_pie_chart(summary_ws, sorted_df["Risk"].value_counts().to_dict(), chart_anchor="I8", data_start_row=15)

    wb.save(normal_path)
    wb.close()
    print(f"  Saved: {normal_path}")

    # ── Output 2: TextJoin report (IPs merged per vulnerability) ─────────
    tj_df = text_join(sorted_df)
    textjoin_path = OUTPUT_DIR / "VA_Server_First_Audit_Report_SCPL_2026_V1_0_TextJoin.xlsx"
    print(f"\nGenerating TextJoin report: {textjoin_path.name}")
    print(f"  TextJoin rows: {len(tj_df)} (from {len(sorted_df)} normal rows)")

    wb2 = load_workbook(TEMPLATE)
    ws_tj = wb2["VA Report"]
    style_va_headers(ws_tj)
    _write_data_rows(ws_tj, tj_df)

    summary_ws2 = wb2["Summary"]
    _write_scope(summary_ws2, tj_df)
    _write_risk_summary(summary_ws2, tj_df, start_row=15)
    build_pie_chart(summary_ws2, tj_df["Risk"].value_counts().to_dict(), chart_anchor="I8", data_start_row=15)

    wb2.save(textjoin_path)
    wb2.close()
    print(f"  Saved: {textjoin_path}")

    print("\nDone! 2 reports saved in output/")


if __name__ == "__main__":
    main()
