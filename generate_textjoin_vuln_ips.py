"""Generate TextJoin Vulnerability Report.

Groups vulnerabilities and joins all host IPs that share the same vulnerability name.
All columns included: Vulnerability Name, Description, Risk, Host, Port, Recommendation, Reference, CVE.

Usage:
    python generate_textjoin_vuln_ips.py RAW_SERVER.xlsx
    python generate_textjoin_vuln_ips.py RAW_SERVER.xlsx -o output/
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

# ── Styling ──────────────────────────────────────────────────────────────
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FONT = Font(name="Cambria", size=11, bold=True)
DATA_FONT = Font(name="Cambria", size=11)

WRAP_COLUMNS = {"Description", "Recommendation", "Reference"}
CENTER_COLUMNS = {"Sr. no", "Risk", "Host", "Port", "CVE"}

TEMPLATE_COLUMNS = [
    "Sr. no",
    "Vulnerability Name",
    "Description",
    "Risk",
    "Host",
    "Port",
    "Recommendation",
    "Reference",
    "CVE",
]

VA_ONLY_RISKS = {"Critical", "High", "Medium", "Low"}

EXCLUDE_NAMES = {
    "SSL Certificate Cannot Be Trusted",
    "SSL Self-Signed Certificate",
}

RISK_WEIGHTS = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}


# ── Load & Clean ─────────────────────────────────────────────────────────
def load_and_clean(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    for col in ["Risk", "Host", "Name"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    df = df.dropna(subset=["Name", "Host"])
    df = df[df["Name"] != ""]
    df = df[df["Host"] != ""]
    return df


def filter_va_only(df: pd.DataFrame) -> pd.DataFrame:
    df = df[df["Risk"].isin(VA_ONLY_RISKS)].copy()
    df = df[~df["Name"].isin(EXCLUDE_NAMES)].copy()
    return df


def exact_dedup(df: pd.DataFrame) -> pd.DataFrame:
    keys = ["Name", "Description", "Risk", "Host"]
    return df.drop_duplicates(subset=keys, keep="first").copy()


# ── Column Mapping ───────────────────────────────────────────────────────
RAW_MAP = {
    "Name": "Vulnerability Name",
    "Description": "Description",
    "Risk": "Risk",
    "Host": "Host",
    "Port": "Port",
    "Solution": "Recommendation",
    "See Also": "Reference",
    "CVE": "CVE",
}


def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    mapped = pd.DataFrame()
    for raw_col, tpl_col in RAW_MAP.items():
        mapped[tpl_col] = df[raw_col] if raw_col in df.columns else ""
    for col in ["Reference", "CVE"]:
        mapped[col] = mapped[col].fillna("N/A").replace("", "N/A")
    mapped["Port"] = pd.to_numeric(mapped["Port"], errors="coerce").fillna(0).astype(int)
    return mapped


# ── TextJoin: group by vulnerability, join IPs ───────────────────────────
def text_join_vuln_ips(df: pd.DataFrame) -> pd.DataFrame:
    """Group rows by Vulnerability Name and join unique Host IPs with comma separator.

    All other columns take the first non-null value from the group.
    """
    join_cols = ["Description", "Risk", "Port", "Recommendation", "Reference", "CVE"]

    def first_non_null(s):
        non_null = s.dropna()
        return non_null.iloc[0] if not non_null.empty else ""

    def unique_hosts(hosts):
        seen = []
        for h in hosts.dropna():
            if h not in seen:
                seen.append(h)
        return ", ".join(seen) if seen else ""

    agg = {}
    for col in join_cols:
        if col in df.columns:
            agg[col] = first_non_null
    if "Host" in df.columns:
        agg["Host"] = unique_hosts

    grouped = df.groupby("Vulnerability Name", sort=False).agg(agg).reset_index()

    grouped["_rw"] = grouped["Risk"].map(RISK_WEIGHTS)
    grouped = grouped.sort_values(by=["_rw"], kind="stable").drop(columns=["_rw"])
    grouped = grouped.reset_index(drop=True)
    grouped.insert(0, "Sr. no", range(1, len(grouped) + 1))
    return grouped


# ── Write to Excel ───────────────────────────────────────────────────────
def _apply_style(cell, col_name: str):
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    if col_name == "Host":
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
    elif col_name in WRAP_COLUMNS:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    elif col_name in CENTER_COLUMNS:
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        cell.alignment = Alignment(vertical="top")


def write_report(data_df: pd.DataFrame, output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "TextJoin Report"

    # Headers
    for col_idx, header in enumerate(TEMPLATE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for i, (_, row) in enumerate(data_df.iterrows()):
        excel_row = i + 2
        for j, col_name in enumerate(TEMPLATE_COLUMNS):
            cell = ws.cell(row=excel_row, column=j + 1)
            value = row.get(col_name)
            if pd.isna(value) or value == "" or value is None:
                cell.value = "N/A"
            else:
                cell.value = value
            _apply_style(cell, col_name)

    # Column widths
    ws.column_dimensions["A"].width = 8    # Sr. no
    ws.column_dimensions["B"].width = 45   # Vulnerability Name
    ws.column_dimensions["C"].width = 40   # Description
    ws.column_dimensions["D"].width = 12   # Risk
    ws.column_dimensions["E"].width = 50   # Host (IPs)
    ws.column_dimensions["F"].width = 10   # Port
    ws.column_dimensions["G"].width = 40   # Recommendation
    ws.column_dimensions["H"].width = 30   # Reference
    ws.column_dimensions["I"].width = 25   # CVE

    wb.save(output_path)
    wb.close()
    return output_path


# ── Main ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Generate TextJoin report (groups IPs by vulnerability, all columns)"
    )
    parser.add_argument("raw_file", type=Path, help="Path to raw Nessus export (.xlsx)")
    parser.add_argument(
        "-o", "--output-dir", type=Path, default=Path("output"),
        help="Output directory (default: output/)",
    )
    parser.add_argument(
        "--output-name", type=str, default=None,
        help="Output filename (without extension). Defaults to raw file name.",
    )
    args = parser.parse_args()

    if not args.raw_file.exists():
        print(f"Error: Raw file not found: {args.raw_file}", file=sys.stderr)
        return 1

    args.output_dir.mkdir(parents=True, exist_ok=True)

    base_name = args.output_name or args.raw_file.stem
    if base_name.endswith(".xlsx"):
        base_name = base_name[:-5]
    output_path = args.output_dir / f"{base_name}_TextJoin_Report.xlsx"

    print("Loading raw data...")
    raw = load_and_clean(args.raw_file)
    print(f"  Raw rows: {len(raw)}")

    print("Filtering to VA risks only...")
    va_only = filter_va_only(raw)
    print(f"  After filter: {len(va_only)}")

    print("Exact dedup...")
    s1 = exact_dedup(va_only)
    print(f"  After dedup: {len(s1)}")

    print("Name+Host dedup...")
    s1b = s1.drop_duplicates(subset=["Name", "Host"], keep="first").copy()
    print(f"  After Name+Host dedup: {len(s1b)}")

    print("Mapping columns...")
    mapped = map_columns(s1b)

    print("TextJoin: grouping vulnerabilities and joining IPs...")
    tj_df = text_join_vuln_ips(mapped)
    print(f"  Unique vulnerabilities: {len(tj_df)}")
    print(f"  Risk breakdown: {tj_df['Risk'].value_counts().to_dict()}")

    print(f"Writing report: {output_path.name}")
    write_report(tj_df, output_path)
    print(f"  Saved: {output_path}")

    print("\nDone!")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
