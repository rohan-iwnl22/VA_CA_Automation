#!/usr/bin/env python3
"""
CA Report - Reverse TEXTJOIN Tool
=================================

Takes a "TextJoin" style CA Audit Report (.xlsx) - where each compliance
finding appears as ONE row with all affected hosts joined into a single
comma-separated "Host" cell - and expands it back into the normal
per-host format (one row per finding + host pair).

HOW TO RUN
----------
1. Make sure Python 3 is installed (python.org) - version 3.8 or newer.
2. Install the one required library (only needed once):
        pip install openpyxl
3. Double-click this file, or run from a terminal:
        python CA_TextJoin_Reverser.py
4. In the window that opens, click "Browse..." to choose your TextJoin
   .xlsx file, confirm/adjust the Save-As location, then click "Convert".

Or run from CLI:
        python CA_TextJoin_Reverser.py input.xlsx -o output.xlsx
"""

import copy
import os
import sys
import threading
import traceback
from collections import Counter

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: The 'openpyxl' package is required.\n"
          "Install it with:  pip install openpyxl")
    sys.exit(1)

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

# ----------------------------------------------------------------------
# Conversion logic
# ----------------------------------------------------------------------

HEADER_ROW = 13
FIRST_DATA_ROW = 14
SHEET_NAME = "CA_Report"
SUMMARY_SHEET = "Summary"

COL_SR = 1
COL_TITLE = 2
COL_HOST = 3
COL_DESC = 4
COL_SOLUTION = 5
COL_STATUS = 6
LAST_COL = 6

STATUS_ORDER = {"FAILED": 0, "WARNING": 1, "PASSED": 2}


def _find_last_data_row(ws):
    last = FIRST_DATA_ROW - 1
    r = FIRST_DATA_ROW
    empty_streak = 0
    max_row = ws.max_row or FIRST_DATA_ROW
    while r <= max_row:
        title = ws.cell(r, COL_TITLE).value
        if title not in (None, ""):
            last = r
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak > 5:
                break
        r += 1
    return last


def _copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.number_format = copy.copy(src_cell.number_format)
    dst_cell.protection = copy.copy(src_cell.protection)
    dst_cell.alignment = copy.copy(src_cell.alignment)


def convert(input_path, output_path, progress_cb=None):
    def report(msg):
        if progress_cb:
            progress_cb(msg)

    report("Opening workbook...")
    wb = load_workbook(input_path, data_only=False)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in the uploaded file.")

    ws = wb[SHEET_NAME]
    last_row = _find_last_data_row(ws)
    if last_row < FIRST_DATA_ROW:
        raise ValueError("No finding rows found to expand.")

    report(f"Reading {last_row - FIRST_DATA_ROW + 1} joined rows...")

    records = []
    for r in range(FIRST_DATA_ROW, last_row + 1):
        title = ws.cell(r, COL_TITLE).value
        if title in (None, ""):
            continue
        host_raw = ws.cell(r, COL_HOST).value
        desc = ws.cell(r, COL_DESC).value
        solution = ws.cell(r, COL_SOLUTION).value
        status = ws.cell(r, COL_STATUS).value

        hosts = []
        if host_raw is not None:
            for h in str(host_raw).split(","):
                h = h.strip()
                if h:
                    hosts.append(h)
        if not hosts:
            hosts = [host_raw]

        for h in hosts:
            records.append({
                "title": title, "host": h, "desc": desc,
                "solution": solution, "status": status,
            })

    report(f"Expanded to {len(records)} rows. Sorting...")

    def sort_key(rec):
        host = rec["host"] if rec["host"] is not None else ""
        status_rank = STATUS_ORDER.get(rec["status"], 99)
        title = rec["title"] if rec["title"] is not None else ""
        return (str(host), status_rank, str(title).lower())

    records.sort(key=sort_key)

    for i, rec in enumerate(records, start=1):
        rec["sr"] = i

    old_count = last_row - FIRST_DATA_ROW + 1
    new_count = len(records)
    template_row = FIRST_DATA_ROW

    if new_count > old_count:
        n_extra = new_count - old_count
        report(f"Inserting {n_extra} extra rows...")
        insert_at = last_row + 1
        ws.insert_rows(insert_at, amount=n_extra)
        template_height = ws.row_dimensions[template_row].height
        for offset in range(n_extra):
            new_r = insert_at + offset
            ws.row_dimensions[new_r].height = template_height
            for c in range(1, LAST_COL + 1):
                _copy_cell_style(ws.cell(template_row, c), ws.cell(new_r, c))
    elif new_count < old_count:
        n_remove = old_count - new_count
        report(f"Removing {n_remove} unused rows...")
        remove_at = FIRST_DATA_ROW + new_count
        ws.delete_rows(remove_at, amount=n_remove)

    report("Writing expanded rows...")
    for i, rec in enumerate(records):
        r = FIRST_DATA_ROW + i
        ws.cell(r, COL_SR).value = rec["sr"]
        ws.cell(r, COL_TITLE).value = rec["title"]
        ws.cell(r, COL_HOST).value = rec["host"]
        ws.cell(r, COL_DESC).value = rec["desc"]
        ws.cell(r, COL_SOLUTION).value = rec["solution"]
        ws.cell(r, COL_STATUS).value = rec["status"]

    stats = {"old_count": old_count, "new_count": new_count, "status_counts": {}}

    if SUMMARY_SHEET in wb.sheetnames:
        report("Updating Summary sheet counts...")
        sws = wb[SUMMARY_SHEET]
        status_counts = Counter(rec["status"] for rec in records if rec["status"])
        stats["status_counts"] = dict(status_counts)
        for r in range(1, sws.max_row + 1):
            label = sws.cell(r, 5).value
            if isinstance(label, str) and label.strip() in ("FAILED", "WARNING", "PASSED"):
                sws.cell(r, 6).value = status_counts.get(label.strip(), 0)
            elif isinstance(label, str) and label.strip() == "Grand Total":
                sws.cell(r, 6).value = sum(status_counts.values())

    report("Saving output file...")
    wb.save(output_path)
    report("Done.")
    return stats


# ----------------------------------------------------------------------
# GUI
# ----------------------------------------------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CA Report - Reverse TEXTJOIN Tool")
        self.geometry("640x480")
        self.resizable(False, False)

        pad = {"padx": 12, "pady": 6}

        tk.Label(
            self, text="CA Report — Reverse TEXTJOIN Tool",
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", **pad)

        tk.Label(
            self,
            text="Upload a TextJoin-format CA report and get back the normal,\n"
                 "one-row-per-host format, ready to paste/use directly.",
            justify="left", fg="#444"
        ).pack(anchor="w", padx=12)

        # --- Input file row ---
        frm_in = tk.Frame(self)
        frm_in.pack(fill="x", **pad)
        tk.Label(frm_in, text="TextJoin file:", width=14, anchor="w").pack(side="left")
        self.input_var = tk.StringVar()
        tk.Entry(frm_in, textvariable=self.input_var).pack(side="left", fill="x", expand=True, padx=6)
        tk.Button(frm_in, text="Browse...", command=self.browse_input).pack(side="left")

        # --- Output file row ---
        frm_out = tk.Frame(self)
        frm_out.pack(fill="x", **pad)
        tk.Label(frm_out, text="Save as:", width=14, anchor="w").pack(side="left")
        self.output_var = tk.StringVar()
        tk.Entry(frm_out, textvariable=self.output_var).pack(side="left", fill="x", expand=True, padx=6)
        tk.Button(frm_out, text="Browse...", command=self.browse_output).pack(side="left")

        # --- Convert button ---
        self.convert_btn = tk.Button(
            self, text="Convert", font=("Segoe UI", 11, "bold"),
            bg="#2563eb", fg="white", activebackground="#1d4ed8",
            command=self.start_convert
        )
        self.convert_btn.pack(pady=10)

        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill="x", padx=12)

        # --- Log box ---
        tk.Label(self, text="Log:").pack(anchor="w", padx=12, pady=(10, 0))
        self.log_box = scrolledtext.ScrolledText(self, height=14, state="disabled", font=("Consolas", 9))
        self.log_box.pack(fill="both", expand=True, padx=12, pady=(0, 12))

    def log(self, msg):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")
        self.update_idletasks()

    def browse_input(self):
        path = filedialog.askopenfilename(
            title="Select the TextJoin .xlsx file",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.input_var.set(path)
            base, ext = os.path.splitext(path)
            suggested = base
            for suffix in ("_TextJoin", "TextJoin"):
                if suggested.endswith(suffix):
                    suggested = suggested[: -len(suffix)]
                    break
            if not suggested.endswith("_Normal") and suggested == base:
                suggested = base + "_Normal"
            self.output_var.set(suggested + ext)

    def browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Save expanded report as...",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if path:
            self.output_var.set(path)

    def start_convert(self):
        in_path = self.input_var.get().strip()
        out_path = self.output_var.get().strip()

        if not in_path or not os.path.isfile(in_path):
            messagebox.showerror("Missing file", "Please select a valid TextJoin .xlsx file.")
            return
        if not out_path:
            messagebox.showerror("Missing output", "Please choose where to save the result.")
            return
        if os.path.abspath(in_path) == os.path.abspath(out_path):
            messagebox.showerror(
                "Same file",
                "Please choose a different Save-As location so the original file isn't overwritten."
            )
            return

        self.convert_btn.configure(state="disabled")
        self.progress.start(12)
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

        thread = threading.Thread(target=self._run_convert, args=(in_path, out_path), daemon=True)
        thread.start()

    def _run_convert(self, in_path, out_path):
        try:
            stats = convert(in_path, out_path, progress_cb=lambda m: self.after(0, self.log, m))
            self.after(0, self._on_success, stats, out_path)
        except Exception as e:
            tb = traceback.format_exc()
            self.after(0, self._on_error, str(e), tb)

    def _on_success(self, stats, out_path):
        self.progress.stop()
        self.convert_btn.configure(state="normal")
        sc = stats.get("status_counts", {})
        summary = (
            f"\nExpanded {stats['old_count']} joined rows -> {stats['new_count']} per-host rows.\n"
            f"Status breakdown: FAILED={sc.get('FAILED',0)}  WARNING={sc.get('WARNING',0)}\n"
            f"Saved to:\n{out_path}"
        )
        self.log(summary)
        messagebox.showinfo("Done", f"Conversion complete!\n\nSaved to:\n{out_path}")

    def _on_error(self, msg, tb):
        self.progress.stop()
        self.convert_btn.configure(state="normal")
        self.log(f"\nERROR: {msg}\n{tb}")
        messagebox.showerror("Conversion failed", msg)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="CA Report - Reverse TEXTJOIN Tool")
    parser.add_argument("input", nargs="?", help="Input TextJoin .xlsx file path")
    parser.add_argument("-o", "--output", help="Output file path (default: input_Normal.xlsx)")
    args = parser.parse_args()

    if args.input:
        in_path = args.input
        if args.output:
            out_path = args.output
        else:
            base, ext = os.path.splitext(in_path)
            out_path = base + "_Normal" + ext
        stats = convert(in_path, out_path, progress_cb=print)
        sc = stats.get("status_counts", {})
        print(f"\nExpanded {stats['old_count']} joined rows -> {stats['new_count']} per-host rows.")
        print(f"Status: FAILED={sc.get('FAILED',0)} WARNING={sc.get('WARNING',0)}")
        print(f"Saved to: {out_path}")
    else:
        app = App()
        app.mainloop()
